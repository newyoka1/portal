#!/usr/bin/env python3
"""
Export LD Analysis to Excel - Simple Version
Creates one sheet per Causeway audience file
Voters appearing in multiple audiences will be on multiple sheets
"""

import os
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import argparse
import pymysql.cursors

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
from utils.db import DB_HOST, DB_USER, DB_PASSWORD, DB_PORT

DB_NAME = "nys_voter_tagging"

# Base output directory
BASE_OUTPUT_DIR = Path(r"C:\Users\georg_2r965zq\OneDrive\Desktop\AUDIANCE DATABASE\analytics_output")
BASE_OUTPUT_DIR.mkdir(exist_ok=True, parents=True)

def connect_db():
    return pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        charset="utf8mb4",
        autocommit=True,
        cursorclass=pymysql.cursors.DictCursor if False else pymysql.cursors.Cursor,
    )

def format_header_row(ws, row=1):
    """Format the header row with bold, background color"""
    for cell in ws[row]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def auto_adjust_columns(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_summary_tab(wb, conn, district_type, district_number):
    """Create summary tab with all audiences"""
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws['A1'] = f"{district_type} {district_number} - Audience Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells('A2:D2')

    # Determine which column to filter by
    column_map = {'LD': 'LDName', 'SD': 'SDName', 'CD': 'CDName'}
    column_name = column_map[district_type]

    # Get totals
    query = f"""
        SELECT COUNT(*) as total_voters
        FROM voter_file
        WHERE {column_name} = %s
    """

    with conn.cursor() as cursor:
        cursor.execute(query, (district_number,))
        total = cursor.fetchone()[0]

    ws['A4'] = "Total Voters in LD:"
    ws['B4'] = total
    ws['B4'].number_format = '#,##0'

    # Get matched count
    query = f"""
        SELECT COUNT(*) as matched_voters
        FROM voter_file
        WHERE {column_name} = %s
          AND origin IS NOT NULL
          AND TRIM(origin) != ''
    """

    with conn.cursor() as cursor:
        cursor.execute(query, (district_number,))
        matched = cursor.fetchone()[0]

    ws['A5'] = "Matched Voters:"
    ws['B5'] = matched
    ws['B5'].number_format = '#,##0'

    ws['A6'] = "Unmatched Voters:"
    ws['B6'] = total - matched
    ws['B6'].number_format = '#,##0'

    # Get all individual audience files
    query = f"""
        SELECT
            audience_file,
            COUNT(DISTINCT f.StateVoterId) as voters
        FROM (
            SELECT
                StateVoterId,
                {column_name},
                TRIM(SUBSTRING_INDEX(SUBSTRING_INDEX(origin, ',', numbers.n), ',', -1)) AS audience_file
            FROM voter_file
            JOIN (
                SELECT 1 n UNION ALL SELECT 2 UNION ALL SELECT 3 UNION ALL SELECT 4
                UNION ALL SELECT 5 UNION ALL SELECT 6 UNION ALL SELECT 7 UNION ALL SELECT 8
            ) numbers
            WHERE origin IS NOT NULL
              AND TRIM(origin) != ''
              AND CHAR_LENGTH(origin) - CHAR_LENGTH(REPLACE(origin, ',', '')) >= n - 1
        ) AS split_origins
        JOIN voter_file f ON split_origins.StateVoterId = f.StateVoterId
        WHERE split_origins.{column_name} = %s
        GROUP BY audience_file
        ORDER BY voters DESC
    """

    ws['A8'] = "Causeway Audience File"
    ws['B8'] = "Unique Voters"
    ws['C8'] = "% of District"
    ws['D8'] = "Tab Name"
    format_header_row(ws, 8)

    with conn.cursor() as cursor:
        cursor.execute(query, (district_number,))
        results = cursor.fetchall()

    row = 9
    for audience, voters in results:
        # Shorten tab name
        tab_name = audience.replace('.csv', '').replace('INDV NYS_', '')[:31]
        pct = round(voters * 100.0 / total, 2)

        ws[f'A{row}'] = audience
        ws[f'B{row}'] = voters
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'C{row}'] = pct
        ws[f'C{row}'].number_format = '0.00'
        ws[f'D{row}'] = tab_name
        row += 1

    # Add unmatched row
    ws[f'A{row}'] = "** NO AUDIENCE MATCH **"
    ws[f'B{row}'] = total - matched
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'C{row}'] = round((total - matched) * 100.0 / total, 2)
    ws[f'C{row}'].number_format = '0.00'
    ws[f'D{row}'] = "Unmatched Voters"

    auto_adjust_columns(ws)
    print(f"  OK Summary tab created ({len(results)} unique audience files)")

    return results

def create_ethnicity_tab(wb, conn, district_type, district_number):
    """Create ethnicity comparison tab - skip if census table doesn't exist"""
    
    # Check if ref_census_surnames table exists
    try:
        with conn.cursor() as cursor:
            cursor.execute("SHOW TABLES LIKE 'ref_census_surnames'")
            if not cursor.fetchone():
                print("  Skipping Ethnicity Analysis (ref_census_surnames table not found)")
                return
    except Exception as e:
        print(f"  Skipping Ethnicity Analysis (error checking table: {e})")
        return
    
    ws = wb.create_sheet("Ethnicity Analysis", 1)

    column_map = {'LD': 'LDName', 'SD': 'SDName', 'CD': 'CDName'}
    column_name = column_map[district_type]

    # Title
    ws['A1'] = f"{district_type} {district_number} - Ethnicity Breakdown"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    # Headers
    ws['A3'] = "Ethnicity"
    ws['B3'] = "Matched Voters"
    ws['C3'] = "Matched %"
    ws['D3'] = "Unmatched Voters"
    ws['E3'] = "Unmatched %"
    ws['F3'] = "Difference"
    format_header_row(ws, 3)

    # Get matched ethnicity
    query_matched = f"""
        SELECT
            COALESCE(e.dominant_ethnicity, 'UNKNOWN') AS ethnicity,
            COUNT(*) AS voters,
            ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER(), 2) AS percentage
        FROM voter_file f
        LEFT JOIN ref_census_surnames e ON e.surname = UPPER(f.LastName)
        WHERE f.{column_name} = %s
          AND f.origin IS NOT NULL
          AND TRIM(f.origin) != ''
        GROUP BY e.dominant_ethnicity
        ORDER BY voters DESC
    """

    # Get unmatched ethnicity
    query_unmatched = f"""
        SELECT
            COALESCE(e.dominant_ethnicity, 'UNKNOWN') AS ethnicity,
            COUNT(*) AS voters,
            ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER(), 2) AS percentage
        FROM voter_file f
        LEFT JOIN ref_census_surnames e ON e.surname = UPPER(f.LastName)
        WHERE f.{column_name} = %s
          AND (f.origin IS NULL OR TRIM(f.origin) = '')
        GROUP BY e.dominant_ethnicity
        ORDER BY voters DESC
    """

    with conn.cursor() as cursor:
        cursor.execute(query_matched, (district_number,))
        matched_results = {row[0]: (row[1], float(row[2])) for row in cursor.fetchall()}

        cursor.execute(query_unmatched, (district_number,))
        unmatched_results = {row[0]: (row[1], float(row[2])) for row in cursor.fetchall()}

    # Combine all ethnicities
    all_ethnicities = set(matched_results.keys()) | set(unmatched_results.keys())

    row = 4
    for ethnicity in sorted(all_ethnicities):
        matched_voters, matched_pct = matched_results.get(ethnicity, (0, 0.0))
        unmatched_voters, unmatched_pct = unmatched_results.get(ethnicity, (0, 0.0))
        diff = matched_pct - unmatched_pct

        ws[f'A{row}'] = ethnicity
        ws[f'B{row}'] = matched_voters
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'C{row}'] = matched_pct
        ws[f'C{row}'].number_format = '0.00'
        ws[f'D{row}'] = unmatched_voters
        ws[f'D{row}'].number_format = '#,##0'
        ws[f'E{row}'] = unmatched_pct
        ws[f'E{row}'].number_format = '0.00'
        ws[f'F{row}'] = diff
        ws[f'F{row}'].number_format = '+0.00;-0.00;0.00'

        # Color code difference
        if diff > 0:
            ws[f'F{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif diff < 0:
            ws[f'F{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        row += 1

    auto_adjust_columns(ws)
    print(f"  OK Ethnicity analysis tab created")

def create_audience_tab(wb, conn, district_type, district_number, audience_file, tab_name):
    """Create tab with full voter list for specific audience file"""

    column_map = {'LD': 'LDName', 'SD': 'SDName', 'CD': 'CDName'}
    column_name = column_map[district_type]

    # Sanitize tab name (Excel limit is 31 chars)
    tab_name = tab_name.replace('.csv', '').replace('INDV NYS_', '')[:31]

    ws = wb.create_sheet(tab_name)

    # Title
    ws['A1'] = f"{audience_file}"
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:L1')

    ws['A2'] = f"{district_type} {district_number}"

    # We'll add the total count in row 3 after we get the results
    # Placeholder for now
    ws['A3'] = "Total Voters: "
    ws['A3'].font = Font(bold=True)

    # Headers
    headers = [
        'StateVoterId', 'FirstName', 'MiddleName', 'LastName',
        'Address', 'City', 'ZIP', 'DOB',
        'LD', 'SD', 'CD'
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=5, column=col_num, value=header)

    format_header_row(ws, 5)

    # Get voters - find anyone who has this audience in their origin field
    query = f"""
        SELECT
            f.StateVoterId,
            f.FirstName,
            f.MiddleName,
            f.LastName,
            f.PrimaryAddress1,
            f.PrimaryCity,
            f.PrimaryZip,
            f.DOB,
            f.LDName,
            f.SDName,
            f.CDName
        FROM voter_file f
        WHERE f.{column_name} = %s
          AND (
              f.origin = %s
              OR f.origin LIKE %s
              OR f.origin LIKE %s
              OR f.origin LIKE %s
          )
        ORDER BY f.LastName, f.FirstName
        LIMIT 100000
    """

    # Build LIKE patterns
    pattern_start = f"{audience_file},%"
    pattern_end = f"%,{audience_file}"
    pattern_middle = f"%,{audience_file},%"

    with conn.cursor() as cursor:
        cursor.execute(query, (district_number, audience_file, pattern_start, pattern_end, pattern_middle))
        results = cursor.fetchall()

    if not results:
        ws['A7'] = "No voters found for this audience in this district"
        ws['B3'] = 0
        ws['B3'].number_format = '#,##0'
        return

    # Update total count in header
    ws['B3'] = len(results)
    ws['B3'].number_format = '#,##0'

    # Write data
    row = 6
    for voter_data in results:
        for col_num, value in enumerate(voter_data, 1):
            ws.cell(row=row, column=col_num, value=value)
        row += 1

        # Progress indicator every 10k rows
        if row % 10000 == 0:
            print(f"    ...{row-6:,} voters written")

    auto_adjust_columns(ws)
    print(f"  OK {tab_name}: {len(results):,} voters")

def create_unmatched_tab(wb, conn, district_type, district_number):
    """Create tab with unmatched voters"""

    column_map = {'LD': 'LDName', 'SD': 'SDName', 'CD': 'CDName'}
    column_name = column_map[district_type]

    # Check if ethnicity table exists
    has_ethnicity = False
    try:
        with conn.cursor() as cursor:
            cursor.execute("SHOW TABLES LIKE 'ref_census_surnames'")
            has_ethnicity = cursor.fetchone() is not None
    except Exception:
        pass

    ws = wb.create_sheet("Unmatched Voters")

    # Title
    ws['A1'] = "Unmatched Voters (Not in Any Audience)"
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:L1')

    ws['A2'] = f"{district_type} {district_number}"

    # Placeholder for total count
    ws['A3'] = "Total Voters: "
    ws['A3'].font = Font(bold=True)

    # Headers - adjust based on whether we have ethnicity data
    if has_ethnicity:
        headers = [
            'StateVoterId', 'FirstName', 'MiddleName', 'LastName',
            'Address', 'City', 'ZIP', 'DOB',
            'LD', 'SD', 'CD', 'Ethnicity'
        ]
    else:
        headers = [
            'StateVoterId', 'FirstName', 'MiddleName', 'LastName',
            'Address', 'City', 'ZIP', 'DOB',
            'LD', 'SD', 'CD'
        ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=5, column=col_num, value=header)

    format_header_row(ws, 5)

    # Get unmatched voters - query varies based on ethnicity table availability
    if has_ethnicity:
        query = f"""
            SELECT
                f.StateVoterId,
                f.FirstName,
                f.MiddleName,
                f.LastName,
                f.PrimaryAddress1,
                f.PrimaryCity,
                f.PrimaryZip,
                f.DOB,
                f.LDName,
                f.SDName,
                f.CDName,
                COALESCE(e.dominant_ethnicity, 'UNKNOWN') as ethnicity
            FROM voter_file f
            LEFT JOIN ref_census_surnames e ON e.surname = UPPER(f.LastName)
            WHERE f.{column_name} = %s
              AND (f.origin IS NULL OR TRIM(f.origin) = '')
            ORDER BY f.LastName, f.FirstName
            LIMIT 100000
        """
    else:
        query = f"""
            SELECT
                f.StateVoterId,
                f.FirstName,
                f.MiddleName,
                f.LastName,
                f.PrimaryAddress1,
                f.PrimaryCity,
                f.PrimaryZip,
                f.DOB,
                f.LDName,
                f.SDName,
                f.CDName
            FROM voter_file f
            WHERE f.{column_name} = %s
              AND (f.origin IS NULL OR TRIM(f.origin) = '')
            ORDER BY f.LastName, f.FirstName
            LIMIT 100000
        """

    with conn.cursor() as cursor:
        cursor.execute(query, (district_number,))
        results = cursor.fetchall()

    # Update total count in header
    ws['B3'] = len(results)
    ws['B3'].number_format = '#,##0'

    # Write data
    row = 6
    for voter_data in results:
        for col_num, value in enumerate(voter_data, 1):
            ws.cell(row=row, column=col_num, value=value)
        row += 1

        if row % 5000 == 0:
            print(f"    ...{row-6:,} unmatched voters written")

    auto_adjust_columns(ws)
    print(f"  OK Unmatched Voters: {len(results):,} voters")

def get_output_dir(district_type, district_number):
    """Get the output directory for a specific district, creating subfolder if needed"""
    district_folder = BASE_OUTPUT_DIR / f"{district_type}_{district_number}"
    district_folder.mkdir(exist_ok=True, parents=True)
    return district_folder

def clean_old_files(district_type, district_number, keep_latest=3):
    """Delete old Excel files for this district, keeping only the most recent ones"""

    output_dir = get_output_dir(district_type, district_number)

    # Find all files for this district
    pattern = f"{district_type}_{district_number}_Complete_Analysis_*.xlsx"
    matching_files = sorted(output_dir.glob(pattern), key=lambda x: x.stat().st_mtime, reverse=True)

    if len(matching_files) <= keep_latest:
        print(f"Found {len(matching_files)} existing file(s) for {district_type} {district_number} (keeping all)")
        return

    # Files to delete (keep the newest keep_latest files)
    files_to_delete = matching_files[keep_latest:]

    print(f"\nCleaning up old files for {district_type} {district_number}:")
    print(f"  Found: {len(matching_files)} files")
    print(f"  Keeping: {keep_latest} newest files")
    print(f"  Deleting: {len(files_to_delete)} old files\n")

    for old_file in files_to_delete:
        try:
            old_file.unlink()
            print(f"  Deleted: {old_file.name}")
        except Exception as e:
            print(f"  [!] Could not delete {old_file.name}: {e}")



def create_custom_ethnicity_tab(wb, conn, district_type, district_number):
    """
    Surname-based ethnicity flags: Russian/Slavic, Jewish, Korean, Italian, Irish.
    Shows counts, pct of district, and party breakdown per group.
    """
    column_map = {"LD": "LDName", "SD": "SDName", "CD": "CDName"}
    col = column_map[district_type]

    ws = wb.create_sheet("Ethnicity (Surname)")

    ws["A1"] = f"{district_type} {district_number} - Surname-Based Ethnicity Groups"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:J1")
    ws["A2"] = "Source: ref_custom_surnames lookup (1,235 surnames) * One voter may appear in multiple groups"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A2:J2")

    groups = [
        ("eth_russian",  "Russian / Slavic",      "DDEEFF"),
        ("eth_jewish",   "Jewish (Ashkenazi)",     "FFF9C4"),
        ("eth_korean",   "Korean",                 "E8F5E9"),
        ("eth_italian",  "Italian",                "FCE4D6"),
        ("eth_irish",    "Irish",                  "F3E5F5"),
    ]

    # --- Section 1: Counts ---
    ws["A4"] = "SECTION 1 - Counts & Match Rate"
    ws["A4"].font = Font(bold=True, size=12, color="1F497D")
    ws.merge_cells("A4:F4")

    headers = ["Ethnic Group", "Voters", "% of District", "Democrat", "Republican", "Unaffiliated"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=5, column=ci, value=h)
    format_header_row(ws, 5)

    with conn.cursor() as cur:
        cur.execute(f"SELECT COUNT(*) FROM voter_file WHERE {col} = %s", (district_number,))
        district_total = int(cur.fetchone()[0])

    row = 6
    for flag_col, label, bg in groups:
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT
                    SUM({flag_col}) total,
                    SUM(CASE WHEN {flag_col}=1 AND OfficialParty='Democrat'     THEN 1 ELSE 0 END) dem,
                    SUM(CASE WHEN {flag_col}=1 AND OfficialParty='Republican'   THEN 1 ELSE 0 END) rep,
                    SUM(CASE WHEN {flag_col}=1 AND OfficialParty='Unaffiliated' THEN 1 ELSE 0 END) una
                FROM voter_file
                WHERE {col} = %s
            """, (district_number,))
            r = cur.fetchone()
        n = int(r[0]) if r[0] else 0
        dem = int(r[1]) if r[1] else 0
        rep = int(r[2]) if r[2] else 0
        una = int(r[3]) if r[3] else 0
        pct = n * 100.0 / district_total if district_total else 0

        for ci, v in enumerate([label, n, round(pct, 2), dem, rep, una], 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.fill = fill
            if ci in (2, 4, 5, 6):
                cell.number_format = "#,##0"
            if ci == 3:
                cell.number_format = "0.00"
        row += 1

    # Total district row
    ws.cell(row=row, column=1, value="Total District").font = Font(bold=True)
    ws.cell(row=row, column=2, value=district_total).number_format = "#,##0"
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 2

    # --- Section 2: In Audiences vs Not ---
    ws.cell(row=row, column=1, value="SECTION 2 - Audience Reach by Ethnic Group")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F497D")
    ws.merge_cells(f"A{row}:G{row}")
    row += 1

    hdrs2 = ["Ethnic Group","Total","In Audiences","Match %","Not in Audiences","Miss %","Largest Audience"]
    for ci, h in enumerate(hdrs2, 1):
        ws.cell(row=row, column=ci, value=h)
    format_header_row(ws, row)
    row += 1

    for flag_col, label, bg in groups:
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        with conn.cursor() as cur:
            # Matched vs unmatched
            cur.execute(f"""
                SELECT
                    SUM({flag_col}) total,
                    SUM(CASE WHEN {flag_col}=1 AND origin IS NOT NULL AND TRIM(origin)!='' THEN 1 ELSE 0 END) matched,
                    SUM(CASE WHEN {flag_col}=1 AND (origin IS NULL OR TRIM(origin)='') THEN 1 ELSE 0 END) unmatched
                FROM voter_file WHERE {col} = %s
            """, (district_number,))
            r = cur.fetchone()
            n = int(r[0]) if r[0] else 0
            matched = int(r[1]) if r[1] else 0
            unmatched = int(r[2]) if r[2] else 0
            match_pct = matched * 100.0 / n if n else 0
            miss_pct = unmatched * 100.0 / n if n else 0

            # Largest audience for this group
            cur.execute(f"""
                SELECT
                    TRIM(SUBSTRING_INDEX(SUBSTRING_INDEX(origin, ',', 1), ',', -1)) AS aud,
                    COUNT(*) n
                FROM voter_file
                WHERE {col} = %s AND {flag_col} = 1
                  AND origin IS NOT NULL AND TRIM(origin) != ''
                GROUP BY 1 ORDER BY 2 DESC LIMIT 1
            """, (district_number,))
            top = cur.fetchone()
            top_aud = top[0].replace(".csv","") if top else "N/A"

        vals = [label, n, matched, round(match_pct,1), unmatched, round(miss_pct,1), top_aud]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.fill = fill
            if ci in (2, 3, 5):
                cell.number_format = "#,##0"
            if ci in (4, 6):
                cell.number_format = "0.0"
        row += 1

    auto_adjust_columns(ws)
    print(f"  OK Ethnicity (Surname) tab created")


def create_modeled_ethnicity_tab(wb, conn, district_type, district_number):
    """
    Ethnicity breakdown using ModeledEthnicity (100% coverage).
    Shows: total voters, matched vs unmatched counts, and party split per ethnicity.
    """
    column_map = {"LD": "LDName", "SD": "SDName", "CD": "CDName"}
    col = column_map[district_type]

    ws = wb.create_sheet("Ethnicity (Modeled)")

    # ── Title ────────────────────────────────────────────────────────────────
    ws["A1"] = f"{district_type} {district_number} — Ethnicity Breakdown (ModeledEthnicity)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")
    ws["A2"] = f"Source: ModeledEthnicity column • Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A2:I2")

    # ── Section 1: Matched vs Unmatched by Ethnicity ─────────────────────────
    ws["A4"] = "SECTION 1 — Audience Match Rate by Ethnicity"
    ws["A4"].font = Font(bold=True, size=12, color="1F497D")
    ws.merge_cells("A4:G4")

    headers1 = ["Ethnicity", "Total Voters", "% of District",
                "In Audiences", "Match Rate", "Not in Audiences", "Miss Rate"]
    for ci, h in enumerate(headers1, 1):
        ws.cell(row=5, column=ci, value=h)
    format_header_row(ws, 5)

    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT
                COALESCE(NULLIF(ModeledEthnicity,''), 'Unknown') AS eth,
                COUNT(*) AS total,
                SUM(CASE WHEN origin IS NOT NULL AND TRIM(origin)!='' THEN 1 ELSE 0 END) AS matched,
                SUM(CASE WHEN origin IS NULL OR TRIM(origin)=''  THEN 1 ELSE 0 END) AS unmatched
            FROM voter_file
            WHERE {col} = %s
            GROUP BY 1 ORDER BY 2 DESC
        """, (district_number,))
        rows = cur.fetchall()

    grand_total = sum(int(r[1]) for r in rows)

    # Ethnicity color map
    eth_colors = {
        "White / Caucasian":       "DDEEFF",
        "Hispanic / Latino":       "FFF2CC",
        "Asian / Pacific Islander":"E2EFDA",
        "Black / African American":"FCE4D6",
        "Other / Multi-Racial":    "EDE7F6",
        "Unknown":                 "F2F2F2",
    }

    row = 6
    for r in rows:
        eth = str(r[0]); total = int(r[1]); matched = int(r[2]); unmatched = int(r[3])
        pct_district = total * 100.0 / grand_total if grand_total else 0
        match_rate   = matched * 100.0 / total if total else 0
        miss_rate    = unmatched * 100.0 / total if total else 0
        bg = eth_colors.get(eth, "FFFFFF")
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        vals = [eth, total, round(pct_district, 1), matched,
                round(match_rate, 1), unmatched, round(miss_rate, 1)]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.fill = fill
            if ci in (2, 4, 6):
                cell.number_format = "#,##0"
            if ci in (3, 5, 7):
                cell.number_format = "0.0"
        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=grand_total).number_format = "#,##0"
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 2

    # ── Section 2: Party Registration by Ethnicity ───────────────────────────
    ws.cell(row=row, column=1, value="SECTION 2 — Party Registration by Ethnicity")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F497D")
    ws.merge_cells(f"A{row}:I{row}")
    row += 1

    # Get all party values present
    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT DISTINCT OfficialParty FROM voter_file
            WHERE {col} = %s AND OfficialParty IS NOT NULL
            ORDER BY OfficialParty
        """, (district_number,))
        all_parties = [r[0] for r in cur.fetchall()]

    # Consolidate minor parties into "Other"
    major = {"Democrat", "Republican", "Unaffiliated", "Conservative", "Working Families"}
    party_cols = [p for p in all_parties if p in major] + ["Other"]

    # Header row
    hdr_row = row
    ws.cell(row=hdr_row, column=1, value="Ethnicity")
    for ci, p in enumerate(party_cols, 2):
        ws.cell(row=hdr_row, column=ci, value=p)
    last_col = len(party_cols) + 2
    ws.cell(row=hdr_row, column=last_col, value="Total")
    format_header_row(ws, hdr_row)
    row += 1

    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT
                COALESCE(NULLIF(ModeledEthnicity,''), 'Unknown') AS eth,
                OfficialParty,
                COUNT(*) AS n
            FROM voter_file
            WHERE {col} = %s
            GROUP BY 1, 2
        """, (district_number,))
        party_raw = cur.fetchall()

    from collections import defaultdict
    party_data = defaultdict(lambda: defaultdict(int))
    eth_totals  = defaultdict(int)
    for eth, party, n in party_raw:
        eth = str(eth) if eth else "Unknown"
        n = int(n)
        bucket = party if party in major else "Other"
        party_data[eth][bucket] += n
        eth_totals[eth] += n

    eth_order = [str(r[0]) for r in rows]  # same order as section 1
    for eth in eth_order:
        bg = eth_colors.get(eth, "FFFFFF")
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        ws.cell(row=row, column=1, value=eth).fill = fill
        for ci, p in enumerate(party_cols, 2):
            cell = ws.cell(row=row, column=ci, value=party_data[eth].get(p, 0))
            cell.number_format = "#,##0"
            cell.fill = fill
        ws.cell(row=row, column=last_col, value=eth_totals[eth]).number_format = "#,##0"
        ws.cell(row=row, column=last_col).fill = fill
        row += 1

    auto_adjust_columns(ws)
    print(f"  OK Ethnicity (Modeled) tab created — {grand_total:,} voters, {len(rows)} groups")


def main():
    parser = argparse.ArgumentParser(description='Export district analysis to Excel - one sheet per audience file')
    parser.add_argument('--ld', type=str, help='Legislative District number (e.g., 63 or 063)')
    parser.add_argument('--sd', type=str, help='State Senate District number (e.g., 5 or 05)')
    parser.add_argument('--cd', type=str, help='Congressional District number (e.g., 3 or 03)')
    parser.add_argument('--keep', type=int, default=0, help='Number of recent files to keep (default: 0 - delete all old files)')
    parser.add_argument('--no-clean', action='store_true', help='Skip cleaning old files')
    args = parser.parse_args()

    # Determine which district type was specified
    district_type = None
    district_number = None
    district_number_padded = None  # For filename

    if args.ld:
        district_type = 'LD'
        district_number_padded = args.ld
        district_number = str(int(args.ld))  # Remove leading zeros for DB query
    elif args.sd:
        district_type = 'SD'
        district_number_padded = args.sd
        district_number = str(int(args.sd))  # Remove leading zeros for DB query
    elif args.cd:
        district_type = 'CD'
        district_number_padded = args.cd
        district_number = str(int(args.cd))  # Remove leading zeros for DB query
    else:
        # Default to LD 63
        district_type = 'LD'
        district_number = '63'
        district_number_padded = '063'
        print("No district specified, defaulting to LD 63")

    print(f"\n{'='*80}")
    print(f"  EXPORTING {district_type} {district_number} TO EXCEL - ONE SHEET PER AUDIENCE FILE")
    print(f"  Database query: {district_type}Name = '{district_number}'")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*80}\n")

    # Clean old files unless --no-clean is specified
    if not args.no_clean:
        clean_old_files(district_type, district_number_padded, keep_latest=args.keep)

    conn = connect_db()

    try:
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create summary tab
        print("Creating Summary tab...")
        audience_list = create_summary_tab(wb, conn, district_type, district_number)

        # Create ethnicity tab
        print("Creating Ethnicity Analysis tab...")

        print("Creating Ethnicity (Modeled) tab...")
        create_modeled_ethnicity_tab(wb, conn, district_type, district_number)

        print("Creating Ethnicity (Surname) tab...")
        create_custom_ethnicity_tab(wb, conn, district_type, district_number)
        create_ethnicity_tab(wb, conn, district_type, district_number)

        print(f"\nCreating tabs for {len(audience_list)} unique audience files...")

        # Create tab for each unique audience file
        for i, (audience_file, voters) in enumerate(audience_list, 1):
            tab_name = audience_file.replace('.csv', '').replace('INDV NYS_', '')
            print(f"\n[{i}/{len(audience_list)}] {audience_file}")
            create_audience_tab(wb, conn, district_type, district_number, audience_file, tab_name)

        # Create unmatched voters tab
        print("\nCreating Unmatched Voters tab...")
        create_unmatched_tab(wb, conn, district_type, district_number)

        # Get output directory and save workbook (use padded number for filename)
        output_dir = get_output_dir(district_type, district_number_padded)
        output_file = output_dir / f"{district_type}_{district_number_padded}_Complete_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print(f"\nSaving workbook...")
        wb.save(output_file)

        print(f"\n{'='*80}")
        print(f"  SUCCESS!")
        print(f"{'='*80}")
        print(f"\nFile saved to:")
        print(f"  {output_file}")
        print(f"\nWorkbook contains:")
        print(f"  - Summary tab")
        print(f"  - Ethnicity Analysis tab (census)")
        print(f"  - Ethnicity (Modeled) tab")
        print(f"  - Ethnicity (Surname) tab")
        print(f"  - {len(audience_list)} unique audience file tabs")
        print(f"  - Unmatched Voters tab")
        print(f"\nTotal tabs: {len(wb.worksheets)}")
        print(f"\nNOTE: Voters appearing in multiple audiences will be listed on multiple sheets")
        print()

    finally:
        conn.close()

if __name__ == "__main__":
    main()