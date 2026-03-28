#!/usr/bin/env python3
"""
Export LD 63 Complete Analysis to Excel
Creates multi-tab Excel workbook with:
- Summary tab (all audiences)
- Ethnicity comparison tab
- Separate tab for each audience with full voter list
"""

import os
import sys
import pymysql
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# MySQL Config
DB_NAME = "NYS_VOTER_TAGGING"
if not MYSQL_PASSWORD:
    raise ValueError("MYSQL_PASSWORD environment variable is required")

# Output directory
OUTPUT_DIR = Path(r"C:\Users\georg_2r965zq\OneDrive\Desktop\AUDIANCE DATABASE\analytics_output")
OUTPUT_DIR.mkdir(exist_ok=True, parents=True)

def connect_db():
    return pymysql.connect(
        host=MYSQL_HOST,
        port=MYSQL_PORT,
        user=MYSQL_USER,
        password=MYSQL_PASSWORD,
        database=DB_NAME,
        charset="utf8mb4",
        autocommit=True,
    )

def format_header_row(ws, row=1):
    """Format the header row with bold, background color"""
    for cell in ws[row]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def auto_adjust_columns(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_summary_tab(wb, conn, ld):
    """Create summary tab with all audiences"""
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws['A1'] = f"LD {ld} - Audience Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells('A2:D2')

    # Get totals
    query = """
        SELECT COUNT(*) as total_voters
        FROM voter_file
        WHERE LDName = %s
    """

    with conn.cursor() as cursor:
        cursor.execute(query, (ld,))
        total = cursor.fetchone()[0]

    ws['A4'] = "Total Voters in LD:"
    ws['B4'] = total
    ws['B4'].number_format = '#,##0'

    # Get matched count
    query = """
        SELECT COUNT(*) as matched_voters
        FROM voter_file
        WHERE LDName = %s
          AND origin IS NOT NULL
          AND TRIM(origin) != ''
    """

    with conn.cursor() as cursor:
        cursor.execute(query, (ld,))
        matched = cursor.fetchone()[0]

    ws['A5'] = "Matched Voters:"
    ws['B5'] = matched
    ws['B5'].number_format = '#,##0'

    ws['A6'] = "Unmatched Voters:"
    ws['B6'] = total - matched
    ws['B6'].number_format = '#,##0'

    # Audience list
    ws['A8'] = "Audience"
    ws['B8'] = "Voters"
    ws['C8'] = "% of District"
    ws['D8'] = "Tab Name"
    format_header_row(ws, 8)

    # Get all audiences
    query = """
        SELECT
            f.origin as audience,
            COUNT(*) as voters,
            ROUND(COUNT(*) * 100.0 / %s, 2) as percentage
        FROM voter_file f
        WHERE f.LDName = %s
          AND f.origin IS NOT NULL
          AND TRIM(f.origin) != ''
        GROUP BY f.origin
        ORDER BY voters DESC
    """

    with conn.cursor() as cursor:
        cursor.execute(query, (total, ld))
        results = cursor.fetchall()

    row = 9
    for aud, voters, pct in results:
        # Shorten tab name
        tab_name = aud.replace('.csv', '').replace('INDV NYS_', '')[:31]

        ws[f'A{row}'] = aud
        ws[f'B{row}'] = voters
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'C{row}'] = pct
        ws[f'C{row}'].number_format = '0.00'
        ws[f'D{row}'] = tab_name
        row += 1

    # Add unmatched row
    ws[f'A{row}'] = "** NO AUDIENCE MATCH **"
    ws[f'B{row}'] = total - matched
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'C{row}'] = round((total - matched) * 100.0 / total, 2)
    ws[f'C{row}'].number_format = '0.00'
    ws[f'D{row}'] = "Unmatched Voters"

    auto_adjust_columns(ws)
    print(f"  OK Summary tab created ({len(results)} audiences)")

def create_ethnicity_tab(wb, conn, ld):
    """Create ethnicity comparison tab"""
    ws = wb.create_sheet("Ethnicity Analysis", 1)

    # Title
    ws['A1'] = f"LD {ld} - Ethnicity Breakdown"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    # Headers
    ws['A3'] = "Ethnicity"
    ws['B3'] = "Matched Voters"
    ws['C3'] = "Matched %"
    ws['D3'] = "Unmatched Voters"
    ws['E3'] = "Unmatched %"
    ws['F3'] = "Difference"
    format_header_row(ws, 3)

    # Get matched ethnicity
    query_matched = """
        SELECT
            COALESCE(e.dominant_ethnicity, 'UNKNOWN') AS ethnicity,
            COUNT(*) AS voters,
            ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER(), 2) AS percentage
        FROM voter_file f
        LEFT JOIN ref_census_surnames e ON e.surname = UPPER(f.LastName)
        WHERE f.LDName = %s
          AND f.origin IS NOT NULL
          AND TRIM(f.origin) != ''
        GROUP BY e.dominant_ethnicity
        ORDER BY voters DESC
    """

    # Get unmatched ethnicity
    query_unmatched = """
        SELECT
            COALESCE(e.dominant_ethnicity, 'UNKNOWN') AS ethnicity,
            COUNT(*) AS voters,
            ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER(), 2) AS percentage
        FROM voter_file f
        LEFT JOIN ref_census_surnames e ON e.surname = UPPER(f.LastName)
        WHERE f.LDName = %s
          AND (f.origin IS NULL OR TRIM(f.origin) = '')
        GROUP BY e.dominant_ethnicity
        ORDER BY voters DESC
    """

    with conn.cursor() as cursor:
        cursor.execute(query_matched, (ld,))
        matched_results = {row[0]: (row[1], float(row[2])) for row in cursor.fetchall()}

        cursor.execute(query_unmatched, (ld,))
        unmatched_results = {row[0]: (row[1], float(row[2])) for row in cursor.fetchall()}

    # Combine all ethnicities
    all_ethnicities = set(matched_results.keys()) | set(unmatched_results.keys())

    row = 4
    for ethnicity in sorted(all_ethnicities):
        matched_voters, matched_pct = matched_results.get(ethnicity, (0, 0.0))
        unmatched_voters, unmatched_pct = unmatched_results.get(ethnicity, (0, 0.0))
        diff = matched_pct - unmatched_pct

        ws[f'A{row}'] = ethnicity
        ws[f'B{row}'] = matched_voters
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'C{row}'] = matched_pct
        ws[f'C{row}'].number_format = '0.00'
        ws[f'D{row}'] = unmatched_voters
        ws[f'D{row}'].number_format = '#,##0'
        ws[f'E{row}'] = unmatched_pct
        ws[f'E{row}'].number_format = '0.00'
        ws[f'F{row}'] = diff
        ws[f'F{row}'].number_format = '+0.00;-0.00;0.00'

        # Color code difference
        if diff > 0:
            ws[f'F{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif diff < 0:
            ws[f'F{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        row += 1

    auto_adjust_columns(ws)
    print(f"  OK Ethnicity analysis tab created")

def create_audience_tab(wb, conn, ld, audience, tab_name):
    """Create tab with full voter list for specific audience"""

    # Sanitize tab name (Excel limit is 31 chars)
    tab_name = tab_name.replace('.csv', '').replace('INDV NYS_', '')[:31]

    ws = wb.create_sheet(tab_name)

    # Title
    ws['A1'] = f"{audience}"
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:L1')

    ws['A2'] = f"LD {ld}"
    ws.merge_cells('A2:L2')

    # Headers
    headers = [
        'StateVoterId', 'FirstName', 'MiddleName', 'LastName',
        'Address', 'City', 'ZIP', 'DOB',
        'LD', 'SD', 'CD', 'Ethnicity'
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=4, column=col_num, value=header)

    format_header_row(ws, 4)

    # Get voters
    query = """
        SELECT
            f.StateVoterId,
            f.FirstName,
            f.MiddleName,
            f.LastName,
            f.PrimaryAddress1,
            f.PrimaryCity,
            f.PrimaryZip,
            f.DOB,
            f.LDName,
            f.SDName,
            f.CDName,
            COALESCE(e.dominant_ethnicity, 'UNKNOWN') as ethnicity
        FROM voter_file f
        LEFT JOIN ref_census_surnames e ON e.surname = UPPER(f.LastName)
        WHERE f.LDName = %s
          AND f.origin = %s
        ORDER BY f.LastName, f.FirstName
        LIMIT 100000
    """

    with conn.cursor() as cursor:
        cursor.execute(query, (ld, audience))
        results = cursor.fetchall()

    if not results:
        ws['A6'] = "No voters found for this audience in this district"
        return

    # Write data
    row = 5
    for voter_data in results:
        for col_num, value in enumerate(voter_data, 1):
            ws.cell(row=row, column=col_num, value=value)
        row += 1

        # Progress indicator every 10k rows
        if row % 10000 == 0:
            print(f"    ...{row-5:,} voters written")

    auto_adjust_columns(ws)
    print(f"  OK {tab_name}: {len(results):,} voters")

def create_unmatched_tab(wb, conn, ld):
    """Create tab with unmatched voters"""

    ws = wb.create_sheet("Unmatched Voters")

    # Title
    ws['A1'] = "Unmatched Voters (Not in Any Audience)"
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:L1')

    ws['A2'] = f"LD {ld}"
    ws.merge_cells('A2:L2')

    # Headers
    headers = [
        'StateVoterId', 'FirstName', 'MiddleName', 'LastName',
        'Address', 'City', 'ZIP', 'DOB',
        'LD', 'SD', 'CD', 'Ethnicity'
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=4, column=col_num, value=header)

    format_header_row(ws, 4)

    # Get unmatched voters
    query = """
        SELECT
            f.StateVoterId,
            f.FirstName,
            f.MiddleName,
            f.LastName,
            f.PrimaryAddress1,
            f.PrimaryCity,
            f.PrimaryZip,
            f.DOB,
            f.LDName,
            f.SDName,
            f.CDName,
            COALESCE(e.dominant_ethnicity, 'UNKNOWN') as ethnicity
        FROM voter_file f
        LEFT JOIN ref_census_surnames e ON e.surname = UPPER(f.LastName)
        WHERE f.LDName = %s
          AND (f.origin IS NULL OR TRIM(f.origin) = '')
        ORDER BY f.LastName, f.FirstName
        LIMIT 100000
    """

    with conn.cursor() as cursor:
        cursor.execute(query, (ld,))
        results = cursor.fetchall()

    # Write data
    row = 5
    for voter_data in results:
        for col_num, value in enumerate(voter_data, 1):
            ws.cell(row=row, column=col_num, value=value)
        row += 1

        if row % 5000 == 0:
            print(f"    ...{row-5:,} unmatched voters written")

    auto_adjust_columns(ws)
    print(f"  OK Unmatched Voters: {len(results):,} voters")

def main():
    import argparse

    parser = argparse.ArgumentParser(description='Export LD analysis to Excel with all audiences')
    parser.add_argument('--ld', type=str, default='63', help='Legislative District number (e.g., 63)')
    args = parser.parse_args()

    ld = args.ld

    print(f"\n{'='*80}")
    print(f"  EXPORTING LD {ld} COMPLETE ANALYSIS TO EXCEL")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*80}\n")

    conn = connect_db()

    try:
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create summary tab
        print("Creating Summary tab...")
        create_summary_tab(wb, conn, ld)

        # Create ethnicity tab
        print("Creating Ethnicity Analysis tab...")
        create_ethnicity_tab(wb, conn, ld)

        # Get all audiences
        query = """
            SELECT DISTINCT origin
            FROM voter_file
            WHERE LDName = %s
              AND origin IS NOT NULL
              AND TRIM(origin) != ''
            ORDER BY origin
        """

        with conn.cursor() as cursor:
            cursor.execute(query, (ld,))
            audiences = [row[0] for row in cursor.fetchall()]

        print(f"\nCreating tabs for {len(audiences)} audiences...")

        # Create tab for each audience
        for i, audience in enumerate(audiences, 1):
            tab_name = audience.replace('.csv', '').replace('INDV NYS_', '')
            print(f"\n[{i}/{len(audiences)}] {audience}")
            create_audience_tab(wb, conn, ld, audience, tab_name)

        # Create unmatched voters tab
        print("\nCreating Unmatched Voters tab...")
        create_unmatched_tab(wb, conn, ld)

        # Save workbook
        output_file = OUTPUT_DIR / f"LD_{ld}_Complete_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print(f"\nSaving workbook...")
        wb.save(output_file)

        print(f"\n{'='*80}")
        print(f"  SUCCESS!")
        print(f"{'='*80}")
        print(f"\nFile saved to:")
        print(f"  {output_file}")
        print(f"\nWorkbook contains:")
        print(f"  - Summary tab")
        print(f"  - Ethnicity Analysis tab")
        print(f"  - {len(audiences)} audience tabs (with full voter lists)")
        print(f"  - Unmatched Voters tab")
        print(f"\nTotal tabs: {len(wb.worksheets)}")
        print()

    finally:
        conn.close()

if __name__ == "__main__":
    main()