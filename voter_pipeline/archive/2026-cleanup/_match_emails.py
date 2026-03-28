"""
Match ProvenDonors list against crm_unified.contacts to find emails.
All rows preserved. Adds email + match_method columns.

Match priority:
  1) State Voter ID (exact)
  2) Phone (any phone field, 10-digit)
  3) Last name + First name + ZIP (cleaned alpha)
  4) First initial + Last name + ZIP
  5) Address number + Street name + ZIP
  6) Last name + Phone (7-digit, no area code)
  7) Nickname expansion + Last name + ZIP
"""
import pymysql
import openpyxl
from dotenv import load_dotenv
import os, re

load_dotenv(r"D:\git\nys-voter-pipeline\.env")

conn = pymysql.connect(
    host=os.getenv("MYSQL_HOST"),
    port=int(os.getenv("MYSQL_PORT")),
    user=os.getenv("MYSQL_USER"),
    password=os.getenv("MYSQL_PASSWORD"),
    database="crm_unified",
    charset="utf8mb4"
)
cur = conn.cursor()

# ── Nickname map (diminutive -> canonical variants) ──
NICKNAMES = {
    'BOB': ['ROBERT'], 'ROB': ['ROBERT'], 'BOBBY': ['ROBERT'],
    'ROBERT': ['BOB', 'ROB', 'BOBBY'],
    'BILL': ['WILLIAM'], 'WILL': ['WILLIAM'], 'BILLY': ['WILLIAM'], 'WILLY': ['WILLIAM'],
    'WILLIAM': ['BILL', 'WILL', 'BILLY', 'WILLY'],
    'JIM': ['JAMES'], 'JIMMY': ['JAMES'], 'JAMIE': ['JAMES'],
    'JAMES': ['JIM', 'JIMMY', 'JAMIE'],
    'MIKE': ['MICHAEL'], 'MIKEY': ['MICHAEL'],
    'MICHAEL': ['MIKE', 'MIKEY'],
    'DICK': ['RICHARD'], 'RICK': ['RICHARD'], 'RICH': ['RICHARD'], 'RICKY': ['RICHARD'],
    'RICHARD': ['DICK', 'RICK', 'RICH', 'RICKY'],
    'TOM': ['THOMAS'], 'TOMMY': ['THOMAS'],
    'THOMAS': ['TOM', 'TOMMY'],
    'TONY': ['ANTHONY'],
    'ANTHONY': ['TONY'],
    'STEVE': ['STEPHEN', 'STEVEN'], 'STEVEN': ['STEVE', 'STEPHEN'], 'STEPHEN': ['STEVE', 'STEVEN'],
    'DAVE': ['DAVID'], 'DAVID': ['DAVE'],
    'DAN': ['DANIEL'], 'DANNY': ['DANIEL'], 'DANIEL': ['DAN', 'DANNY'],
    'JOE': ['JOSEPH'], 'JOEY': ['JOSEPH'], 'JOSEPH': ['JOE', 'JOEY'],
    'MATT': ['MATTHEW'], 'MATTHEW': ['MATT'],
    'CHRIS': ['CHRISTOPHER', 'CHRISTINE', 'CHRISTINA'],
    'CHRISTOPHER': ['CHRIS'], 'CHRISTINE': ['CHRIS'], 'CHRISTINA': ['CHRIS'],
    'PAT': ['PATRICK', 'PATRICIA'], 'PATRICK': ['PAT'], 'PATRICIA': ['PAT'],
    'ED': ['EDWARD', 'EDWIN', 'EDMUND'], 'EDDIE': ['EDWARD'],
    'EDWARD': ['ED', 'EDDIE'], 'EDWIN': ['ED'],
    'JOHN': ['JACK', 'JOHNNY'], 'JACK': ['JOHN'], 'JOHNNY': ['JOHN'],
    'CHARLIE': ['CHARLES'], 'CHUCK': ['CHARLES'], 'CHARLES': ['CHARLIE', 'CHUCK'],
    'BETTY': ['ELIZABETH'], 'BETH': ['ELIZABETH'], 'LIZ': ['ELIZABETH'], 'LIZZY': ['ELIZABETH'],
    'ELIZABETH': ['BETTY', 'BETH', 'LIZ', 'LIZZY'],
    'PEGGY': ['MARGARET'], 'MEG': ['MARGARET'], 'MAGGIE': ['MARGARET'],
    'MARGARET': ['PEGGY', 'MEG', 'MAGGIE'],
    'KATE': ['KATHERINE', 'KATHLEEN', 'CATHERINE'],
    'KATHERINE': ['KATE', 'KATHY'], 'KATHLEEN': ['KATE', 'KATHY'], 'CATHERINE': ['KATE', 'CATHY'],
    'KATHY': ['KATHERINE', 'KATHLEEN'], 'CATHY': ['CATHERINE'],
    'SUE': ['SUSAN', 'SUZANNE'], 'SUSAN': ['SUE'], 'SUZANNE': ['SUE'],
    'JENNY': ['JENNIFER'], 'JEN': ['JENNIFER'], 'JENNIFER': ['JENNY', 'JEN'],
    'LARRY': ['LAWRENCE'], 'LAWRENCE': ['LARRY'],
    'FRANK': ['FRANCIS', 'FRANKLIN'], 'FRANCIS': ['FRANK'], 'FRANKLIN': ['FRANK'],
    'FRED': ['FREDERICK'], 'FREDERICK': ['FRED'],
    'AL': ['ALBERT', 'ALAN', 'ALLEN'], 'ALBERT': ['AL'], 'ALAN': ['AL'], 'ALLEN': ['AL'],
    'ALEX': ['ALEXANDER', 'ALEXANDRA'], 'ALEXANDER': ['ALEX'], 'ALEXANDRA': ['ALEX'],
    'TED': ['THEODORE', 'EDWARD'], 'THEODORE': ['TED'],
    'RAY': ['RAYMOND'], 'RAYMOND': ['RAY'],
    'HARRY': ['HAROLD', 'HENRY'], 'HAROLD': ['HARRY'],
    'PHIL': ['PHILIP', 'PHILLIP'], 'PHILIP': ['PHIL'], 'PHILLIP': ['PHIL'],
    'RON': ['RONALD'], 'RONALD': ['RON'],
    'JERRY': ['GERALD', 'JEROME'], 'GERALD': ['JERRY'], 'JEROME': ['JERRY'],
    'DEBBIE': ['DEBORAH'], 'DEB': ['DEBORAH'], 'DEBORAH': ['DEBBIE', 'DEB'],
    'SANDY': ['SANDRA'], 'SANDRA': ['SANDY'],
    'BARB': ['BARBARA'], 'BARBARA': ['BARB'],
    'DONNA': ['DAWN'],
    'ANN': ['ANNE', 'ANNA'], 'ANNE': ['ANN', 'ANNA'], 'ANNA': ['ANN', 'ANNE'],
    'NICK': ['NICHOLAS'], 'NICHOLAS': ['NICK'],
    'DOUG': ['DOUGLAS'], 'DOUGLAS': ['DOUG'],
    'KEN': ['KENNETH'], 'KENNETH': ['KEN'],
    'DON': ['DONALD'], 'DONALD': ['DON'],
    'GREG': ['GREGORY'], 'GREGORY': ['GREG'],
    'JEFF': ['JEFFREY'], 'JEFFREY': ['JEFF'],
    'ANDY': ['ANDREW'], 'DREW': ['ANDREW'], 'ANDREW': ['ANDY', 'DREW'],
    'PETE': ['PETER'], 'PETER': ['PETE'],
    'MARK': ['MARC'], 'MARC': ['MARK'],
    'JON': ['JONATHAN'], 'JONATHAN': ['JON'],
    'SAM': ['SAMUEL', 'SAMANTHA'], 'SAMUEL': ['SAM'], 'SAMANTHA': ['SAM'],
    'BEN': ['BENJAMIN'], 'BENJAMIN': ['BEN'],
    'TIM': ['TIMOTHY'], 'TIMOTHY': ['TIM'],
}

def digits_only(s):
    if not s: return None
    d = re.sub(r'\D', '', str(s))
    return d[-10:] if len(d) >= 10 else None

def last7(s):
    if not s: return None
    d = re.sub(r'\D', '', str(s))
    return d[-7:] if len(d) >= 7 else None

def clean_alpha(s):
    if not s: return None
    return re.sub(r'[^A-Z]', '', str(s).upper())

def parse_street(addr):
    """Extract (number, street_name_alpha) from address string."""
    if not addr: return None, None
    addr = str(addr).upper().strip()
    m = re.match(r'^(\d+)\s+(.+)', addr)
    if not m: return None, None
    num = m.group(1)
    street = re.sub(r'[^A-Z]', '', m.group(2).split(' APT')[0].split(' UNIT')[0].split(' #')[0].split(' STE')[0])
    return num, street if street else (num, None)

# ── Load CRM contacts with at least one email ──
print("Loading CRM contacts with emails...")
cur.execute("""
    SELECT vf_state_voter_id,
           phone_1, phone_2, phone_3, phone_4, phone_5, mobile,
           clean_first, clean_last, zip5,
           email_1, email_2, email_3, email_4, email_5,
           address, vf_address
    FROM contacts
    WHERE COALESCE(email_1, email_2, email_3, email_4, email_5) IS NOT NULL
      AND COALESCE(email_1, email_2, email_3, email_4, email_5) != ''
""")
rows = cur.fetchall()
print(f"  CRM contacts with email: {len(rows):,}")

# Build lookup indexes
by_voter_id = {}
by_phone = {}
by_name_zip = {}
by_initial_last_zip = {}
by_addr_zip = {}
by_last_phone7 = {}
by_nick_last_zip = {}

for row in rows:
    (vid, p1, p2, p3, p4, p5, mob, cfirst, clast, z5,
     e1, e2, e3, e4, e5, addr, vf_addr) = row
    best_email = e1 or e2 or e3 or e4 or e5
    if not best_email:
        continue

    cf = clean_alpha(cfirst)
    cl = clean_alpha(clast)
    z5s = z5.strip() if z5 else None

    # 1) Voter ID
    if vid and vid.strip():
        by_voter_id[vid.strip()] = best_email

    # 2) Phone (10-digit)
    all_phones = []
    for ph in [p1, p2, p3, p4, p5, mob]:
        d = digits_only(ph)
        if d:
            all_phones.append(d)
            if d not in by_phone:
                by_phone[d] = best_email

    # 3) Full name + zip
    if cf and cl and z5s:
        key = (cf, cl, z5s)
        if key not in by_name_zip:
            by_name_zip[key] = best_email

    # 4) First initial + last + zip
    if cf and cl and z5s:
        key = (cf[0], cl, z5s)
        if key not in by_initial_last_zip:
            by_initial_last_zip[key] = best_email

    # 5) Address number + street + zip
    for a in [addr, vf_addr]:
        num, street = parse_street(a)
        if num and street and z5s:
            key = (num, street, z5s)
            if key not in by_addr_zip:
                by_addr_zip[key] = best_email

    # 6) Last name + phone last 7
    if cl:
        for ph in [p1, p2, p3, p4, p5, mob]:
            d7 = last7(ph)
            if d7:
                key = (cl, d7)
                if key not in by_last_phone7:
                    by_last_phone7[key] = best_email

    # 7) Nickname variants + last + zip
    if cf and cl and z5s:
        variants = NICKNAMES.get(cf, [])
        for nick in variants:
            key = (nick, cl, z5s)
            if key not in by_nick_last_zip:
                by_nick_last_zip[key] = best_email

print(f"  Voter ID index:        {len(by_voter_id):,}")
print(f"  Phone index:           {len(by_phone):,}")
print(f"  Name+ZIP index:        {len(by_name_zip):,}")
print(f"  Initial+Last+ZIP idx:  {len(by_initial_last_zip):,}")
print(f"  Address+ZIP index:     {len(by_addr_zip):,}")
print(f"  Last+Phone7 index:     {len(by_last_phone7):,}")
print(f"  Nickname+Last+ZIP idx: {len(by_nick_last_zip):,}")

conn.close()

# ── Load Excel file ──
print("\nLoading ProvenDonors Excel...")
src = r"C:\Users\georg_2r965zq\OneDrive\Desktop\ProvenDonors_CList_10723hh_20260218.xlsx"
wb = openpyxl.load_workbook(src)
ws = wb.active

headers = [c.value for c in ws[1]]
print(f"  Rows: {ws.max_row - 1:,}")

def col_idx(name):
    return headers.index(name)

idx_sboeid = col_idx('sboeid')
idx_phone = col_idx('phonedigits')
idx_cell = col_idx('cellphone')
idx_first = col_idx('FIRSTNAME')
idx_last = col_idx('LASTNAME')
idx_zip = col_idx('ZIPCODE')
idx_email = col_idx('email')
idx_addr = col_idx('ADDRESSLINE1')
idx_pnum = col_idx('PRIMARYNUMBER')
idx_street = col_idx('STREETNAME')

# Add match_method header
method_col = ws.max_column + 1
ws.cell(row=1, column=method_col, value='match_method')

# ── Match ──
print("\nMatching...")
counts = {'voter_id': 0, 'phone': 0, 'name_zip': 0,
           'initial_zip': 0, 'addr_zip': 0, 'last_phone7': 0, 'nickname_zip': 0}
total = 0

for row in ws.iter_rows(min_row=2, values_only=False):
    total += 1
    vals = [c.value for c in row]
    email_found = None
    match_method = None

    sboeid = vals[idx_sboeid]
    raw_first = vals[idx_first]
    raw_last = vals[idx_last]
    raw_zip = str(vals[idx_zip]).strip() if vals[idx_zip] else None
    cf = clean_alpha(raw_first)
    cl = clean_alpha(raw_last)

    # 1) State Voter ID
    if sboeid and str(sboeid).strip():
        email_found = by_voter_id.get(str(sboeid).strip())
        if email_found: match_method = 'voter_id'

    # 2) Phone 10-digit
    if not email_found:
        for ph in [vals[idx_phone], vals[idx_cell]]:
            d = digits_only(ph)
            if d:
                email_found = by_phone.get(d)
                if email_found:
                    match_method = 'phone'
                    break

    # 3) Full name + ZIP
    if not email_found and cf and cl and raw_zip:
        email_found = by_name_zip.get((cf, cl, raw_zip))
        if email_found: match_method = 'name_zip'

    # 4) First initial + last + ZIP
    if not email_found and cf and cl and raw_zip:
        email_found = by_initial_last_zip.get((cf[0], cl, raw_zip))
        if email_found: match_method = 'initial_zip'

    # 5) Address number + street + ZIP
    if not email_found and raw_zip:
        pnum = str(vals[idx_pnum]).strip() if vals[idx_pnum] else None
        street = clean_alpha(vals[idx_street])
        if pnum and street:
            email_found = by_addr_zip.get((pnum, street, raw_zip))
            if email_found: match_method = 'addr_zip'
        if not email_found:
            num2, street2 = parse_street(vals[idx_addr])
            if num2 and street2:
                email_found = by_addr_zip.get((num2, street2, raw_zip))
                if email_found: match_method = 'addr_zip'

    # 6) Last name + phone last 7
    if not email_found and cl:
        for ph in [vals[idx_phone], vals[idx_cell]]:
            d7 = last7(ph)
            if d7:
                email_found = by_last_phone7.get((cl, d7))
                if email_found:
                    match_method = 'last_phone7'
                    break

    # 7) Nickname + last + ZIP
    if not email_found and cf and cl and raw_zip:
        variants = NICKNAMES.get(cf, [])
        for nick in variants:
            email_found = by_nick_last_zip.get((nick, cl, raw_zip))
            if email_found:
                match_method = 'nickname_zip'
                break

    if email_found:
        row[idx_email].value = email_found
        ws.cell(row=row[0].row, column=method_col, value=match_method)
        counts[match_method] += 1

# ── Save ──
out_path = r"C:\Users\georg_2r965zq\OneDrive\Desktop\ProvenDonors_WithEmails.xlsx"
wb.save(out_path)

total_matched = sum(counts.values())
print(f"\n{'='*50}")
print(f"RESULTS")
print(f"{'='*50}")
print(f"Total records:              {total:,}")
print(f"  1) Voter ID match:        {counts['voter_id']:,}")
print(f"  2) Phone match:           {counts['phone']:,}")
print(f"  3) Name+ZIP match:        {counts['name_zip']:,}")
print(f"  4) Initial+Last+ZIP:      {counts['initial_zip']:,}")
print(f"  5) Address+ZIP match:     {counts['addr_zip']:,}")
print(f"  6) Last+Phone7 match:     {counts['last_phone7']:,}")
print(f"  7) Nickname+ZIP match:    {counts['nickname_zip']:,}")
print(f"{'='*50}")
print(f"TOTAL with email:           {total_matched:,}  ({100*total_matched/total:.1f}%)")
print(f"Still missing:              {total - total_matched:,}")
print(f"\nSaved to: {out_path}")
