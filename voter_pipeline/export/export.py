#!/usr/bin/env python3
"""
Export LD Analysis to Excel - Simple Version
Creates one sheet per Causeway audience file
Voters appearing in multiple audiences will be on multiple sheets
"""

import os
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import argparse
import pymysql.cursors

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
from utils.db import DB_HOST, DB_USER, DB_PASSWORD, DB_PORT

DB_NAME = "nys_voter_tagging"

# ── District filter helpers ───────────────────────────────────────────────────
# col is None when district_type == 'STATEWIDE' (no filter).
# _dc  → bare condition string  e.g. "LDName=%s"  or "1=1"
# _dp  → param tuple            e.g. ('63',)       or ()
def _dc(col):   return f'{col}=%s' if col else '1=1'
def _dp(col, val): return (val,) if col else ()

_COL_MAP = {'LD': 'LDName', 'SD': 'SDName', 'CD': 'CDName',
            'COUNTY': 'CountyName', 'STATEWIDE': None}

# Base output directory
BASE_OUTPUT_DIR = Path(__file__).parent.parent / "output"
BASE_OUTPUT_DIR.mkdir(exist_ok=True, parents=True)

def connect_db():
    return pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        charset="utf8mb4",
        autocommit=True,
        cursorclass=pymysql.cursors.DictCursor if False else pymysql.cursors.Cursor,
    )

def _has_crm_email(conn):
    """Check if crm_email column exists on voter_file (pre-computed by enrich_voter_crm)."""
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT crm_email FROM voter_file LIMIT 0")
            return True
    except Exception:
        return False

def _has_crm_phone(conn):
    """Check if crm_phone/crm_mobile columns exist on voter_file."""
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT crm_phone, crm_mobile FROM voter_file LIMIT 0")
            return True
    except Exception:
        return False

def format_header_row(ws, row=1):
    """Format the header row with bold, background color"""
    for cell in ws[row]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def auto_adjust_columns(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_summary_tab(wb, conn, district_type, district_number, exclude_turnout=False):
    """Summary tab: Turnout Models section, Issue Audiences section, District Totals.
    Each section shows per-audience unique voter counts plus a deduplicated combined total.
    Set exclude_turnout=True to omit the turnout models section (used by voter-contact export)."""
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_properties.tabColor = "808080"
    col = _COL_MAP[district_type]

    # -- styles ----------------------------------------------------------------
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill   = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    sec_font   = Font(bold=True, size=12)
    sec_fill   = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    sub_fill   = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")
    total_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    pct_fmt    = '0.0%'
    num_fmt    = '#,##0'

    def hrow(ws, row, labels):
        for ci, v in enumerate(labels, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")

    def drow(ws, row, audience, voters, total, fill=None):
        tab = audience.replace('.csv','').replace('INDV NYS_','')[:31]
        pct = voters / total if total else 0
        ws.cell(row=row, column=1, value=audience)
        c = ws.cell(row=row, column=2, value=voters); c.number_format = num_fmt
        c = ws.cell(row=row, column=3, value=pct);    c.number_format = pct_fmt
        ws.cell(row=row, column=4, value=tab)
        if fill:
            for ci in range(1, 5):
                ws.cell(row=row, column=ci).fill = fill

    # -- fetch district totals -------------------------------------------------
    with conn.cursor() as cur:
        cur.execute(f"SELECT COUNT(*) FROM voter_file WHERE {_dc(col)}", _dp(col, district_number))
        total_voters = cur.fetchone()[0]
        cur.execute(f"SELECT COUNT(*) FROM voter_file WHERE {_dc(col)} AND origin IS NOT NULL AND TRIM(origin)!=''",
                    _dp(col, district_number))
        total_matched = cur.fetchone()[0]

    # -- fetch per-audience counts ---------------------------------------------
    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT aud, COUNT(DISTINCT StateVoterId) AS voters
            FROM (
                SELECT StateVoterId,
                    TRIM(SUBSTRING_INDEX(SUBSTRING_INDEX(origin,',',n.n),',',-1)) AS aud
                FROM voter_file
                JOIN (SELECT 1 n UNION SELECT 2 UNION SELECT 3 UNION SELECT 4
                      UNION SELECT 5 UNION SELECT 6 UNION SELECT 7 UNION SELECT 8
                      UNION SELECT 9 UNION SELECT 10 UNION SELECT 11 UNION SELECT 12) n
                WHERE {_dc(col)} AND origin IS NOT NULL AND TRIM(origin)!=''
                  AND CHAR_LENGTH(origin)-CHAR_LENGTH(REPLACE(origin,',','')) >= n.n-1
            ) x
            GROUP BY aud
            ORDER BY voters DESC
        """, _dp(col, district_number))
        all_rows = cur.fetchall()  # [(audience_file, voters), ...]

    turnout_prefixes = ('HT ', 'MT ', 'LT ')
    turnout = [(a, v) for a, v in all_rows if a.upper().startswith(turnout_prefixes)]
    issues  = [(a, v) for a, v in all_rows if not a.upper().startswith(turnout_prefixes)]

    # -- deduplicated counts ---------------------------------------------------
    def dedup_count(audiences):
        if not audiences: return 0
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT COUNT(DISTINCT StateVoterId) FROM voter_file
                WHERE {_dc(col)} AND (
                    {' OR '.join([f"FIND_IN_SET(%s, REPLACE(origin, ', ', ','))>0"]*len(audiences))}
                )
            """, (*_dp(col, district_number), *[a for a, _ in audiences]))
            return cur.fetchone()[0]

    turnout_dedup = 0 if exclude_turnout else dedup_count(turnout)
    issues_dedup  = dedup_count(issues)

    # -- build sheet -----------------------------------------------------------
    ws['A1'] = f"{'Statewide' if district_type == 'STATEWIDE' else f'{district_type} {district_number}'} - Audience Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells('A2:D2')

    row = 4

    # -- ISSUE AUDIENCES section -----------------------------------------------
    ws.cell(row=row, column=1, value="ISSUE AUDIENCES").font = sec_font
    for ci in range(1,5): ws.cell(row=row, column=ci).fill = sec_fill
    row += 1
    hrow(ws, row, ["Audience File", "Unique Voters", "% of District", "Tab Name"])
    row += 1
    for aud, voters in issues:
        drow(ws, row, aud, voters, total_voters)
        row += 1
    # combined dedup row
    c = ws.cell(row=row, column=1, value="? COMBINED UNIQUE (no double-count)")
    c.font = Font(bold=True, italic=True)
    c = ws.cell(row=row, column=2, value=issues_dedup); c.number_format = num_fmt; c.font = Font(bold=True)
    c = ws.cell(row=row, column=3, value=issues_dedup/total_voters if total_voters else 0)
    c.number_format = pct_fmt; c.font = Font(bold=True)
    for ci in range(1,5): ws.cell(row=row, column=ci).fill = total_fill
    row += 2

    # -- TURNOUT MODELS section ------------------------------------------------
    if not exclude_turnout:
        ws.cell(row=row, column=1, value="TURNOUT MODELS  (HT / MT / LT)").font = sec_font
        for ci in range(1,5): ws.cell(row=row, column=ci).fill = sec_fill
        row += 1
        hrow(ws, row, ["Audience File", "Unique Voters", "% of District", "Tab Name"])
        row += 1
        for aud, voters in sorted(turnout, key=lambda x: (x[0][:2], -x[1])):
            drow(ws, row, aud, voters, total_voters)
            row += 1
        # combined dedup row
        c = ws.cell(row=row, column=1, value="? COMBINED UNIQUE (no double-count)")
        c.font = Font(bold=True, italic=True)
        c = ws.cell(row=row, column=2, value=turnout_dedup); c.number_format = num_fmt; c.font = Font(bold=True)
        c = ws.cell(row=row, column=3, value=turnout_dedup/total_voters if total_voters else 0)
        c.number_format = pct_fmt; c.font = Font(bold=True)
        for ci in range(1,5): ws.cell(row=row, column=ci).fill = total_fill
        row += 2

    # -- DISTRICT TOTALS -------------------------------------------------------
    ws.cell(row=row, column=1, value="DISTRICT TOTALS").font = sec_font
    for ci in range(1,5): ws.cell(row=row, column=ci).fill = sec_fill
    row += 1
    totals = [
        ("Total Voters in District",  total_voters,              1.0),
        ("Total Matched (unique)",     total_matched,             total_matched/total_voters if total_voters else 0),
        ("Unmatched",                  total_voters-total_matched, (total_voters-total_matched)/total_voters if total_voters else 0),
    ]
    for label, val, pct in totals:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=val); c.number_format = num_fmt
        c = ws.cell(row=row, column=3, value=pct); c.number_format = pct_fmt
        row += 1

    auto_adjust_columns(ws)
    if exclude_turnout:
        print(f"  OK Summary tab created ({len(issues)} issue audiences, turnout excluded)")
    else:
        print(f"  OK Summary tab created ({len(turnout)} turnout models, {len(issues)} issue audiences)")
    return all_rows

def create_audience_tab(wb, conn, district_type, district_number, audience_file, tab_name, tab_color="FFC000"):
    """Create tab with full voter list for specific audience file"""

    column_name = _COL_MAP[district_type]

    # Sanitize tab name (Excel limit is 31 chars)
    tab_name = tab_name.replace('.csv', '').replace('INDV NYS_', '')[:31]

    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = tab_color

    # Title
    ws['A1'] = f"{audience_file}"
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:L1')

    ws['A2'] = 'Statewide' if district_type == 'STATEWIDE' else f"{district_type} {district_number}"

    # We'll add the total count in row 3 after we get the results
    # Placeholder for now
    ws['A3'] = "Total Voters: "
    ws['A3'].font = Font(bold=True)

    # Headers
    headers = [
        'StateVoterId', 'FirstName', 'MiddleName', 'LastName',
        'Address', 'City', 'ZIP', 'DOB',
        'LD', 'SD', 'CD'
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=5, column=col_num, value=header)

    format_header_row(ws, 5)

    # Get voters - find anyone who has this audience in their origin field
    # Statewide uses a higher row limit (Excel max ~1M); district uses 100k
    row_limit = 1000000 if district_type == 'STATEWIDE' else 100000
    dist_cond = f'AND f.{column_name} = %s' if column_name else ''
    query = f"""
        SELECT
            f.StateVoterId,
            f.FirstName,
            f.MiddleName,
            f.LastName,
            f.PrimaryAddress1,
            f.PrimaryCity,
            f.PrimaryZip,
            f.DOB,
            f.LDName,
            f.SDName,
            f.CDName
        FROM voter_file f
        WHERE 1=1 {dist_cond}
          AND (
              f.origin = %s
              OR f.origin LIKE %s
              OR f.origin LIKE %s
              OR f.origin LIKE %s
          )
        ORDER BY f.LastName, f.FirstName
        LIMIT {row_limit}
    """

    # Build LIKE patterns
    pattern_start = f"{audience_file},%"
    pattern_end = f"%,{audience_file}"
    pattern_middle = f"%,{audience_file},%"

    with conn.cursor() as cursor:
        cursor.execute(query, (*_dp(column_name, district_number),
                               audience_file, pattern_start, pattern_end, pattern_middle))
        results = cursor.fetchall()

    if not results:
        ws['A7'] = "No voters found for this audience in this district"
        ws['B3'] = 0
        ws['B3'].number_format = '#,##0'
        return

    # Update total count in header
    ws['B3'] = len(results)
    ws['B3'].number_format = '#,##0'

    # Write data
    row = 6
    for voter_data in results:
        for col_num, value in enumerate(voter_data, 1):
            ws.cell(row=row, column=col_num, value=value)
        row += 1

        # Progress indicator every 10k rows
        if row % 10000 == 0:
            print(f"    ...{row-6:,} voters written")

    auto_adjust_columns(ws)
    print(f"  OK {tab_name}: {len(results):,} voters")

def create_unmatched_tab(wb, conn, district_type, district_number):
    """Create tab with unmatched voters"""

    column_name = _COL_MAP[district_type]

    # Check if ethnicity table exists
    has_ethnicity = False
    try:
        with conn.cursor() as cursor:
            cursor.execute("SHOW TABLES LIKE 'ref_census_surnames'")
            has_ethnicity = cursor.fetchone() is not None
    except Exception:
        pass

    ws = wb.create_sheet("Unmatched Voters")
    ws.sheet_properties.tabColor = "595959"

    # Title
    ws['A1'] = "Unmatched Voters (Not in Any Audience)"
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:L1')

    ws['A2'] = 'Statewide' if district_type == 'STATEWIDE' else f"{district_type} {district_number}"

    # Placeholder for total count
    ws['A3'] = "Total Voters: "
    ws['A3'].font = Font(bold=True)

    # Headers - adjust based on whether we have ethnicity data
    if has_ethnicity:
        headers = [
            'StateVoterId', 'FirstName', 'MiddleName', 'LastName',
            'Address', 'City', 'ZIP', 'DOB',
            'LD', 'SD', 'CD', 'Ethnicity'
        ]
    else:
        headers = [
            'StateVoterId', 'FirstName', 'MiddleName', 'LastName',
            'Address', 'City', 'ZIP', 'DOB',
            'LD', 'SD', 'CD'
        ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=5, column=col_num, value=header)

    format_header_row(ws, 5)

    # Get unmatched voters - query varies based on ethnicity table availability
    dist_cond   = f'AND f.{column_name} = %s' if column_name else ''
    row_limit   = 1000000 if district_type == 'STATEWIDE' else 100000
    if has_ethnicity:
        query = f"""
            SELECT
                f.StateVoterId,
                f.FirstName,
                f.MiddleName,
                f.LastName,
                f.PrimaryAddress1,
                f.PrimaryCity,
                f.PrimaryZip,
                f.DOB,
                f.LDName,
                f.SDName,
                f.CDName,
                COALESCE(e.dominant_ethnicity, 'UNKNOWN') as ethnicity
            FROM voter_file f
            LEFT JOIN ref_census_surnames e ON e.surname = UPPER(f.LastName)
            WHERE 1=1 {dist_cond}
              AND (f.origin IS NULL OR TRIM(f.origin) = '')
            ORDER BY f.LastName, f.FirstName
            LIMIT {row_limit}
        """
    else:
        query = f"""
            SELECT
                f.StateVoterId,
                f.FirstName,
                f.MiddleName,
                f.LastName,
                f.PrimaryAddress1,
                f.PrimaryCity,
                f.PrimaryZip,
                f.DOB,
                f.LDName,
                f.SDName,
                f.CDName
            FROM voter_file f
            WHERE 1=1 {dist_cond}
              AND (f.origin IS NULL OR TRIM(f.origin) = '')
            ORDER BY f.LastName, f.FirstName
            LIMIT {row_limit}
        """

    with conn.cursor() as cursor:
        cursor.execute(query, _dp(column_name, district_number))
        results = cursor.fetchall()

    # Update total count in header
    ws['B3'] = len(results)
    ws['B3'].number_format = '#,##0'

    # Write data
    row = 6
    for voter_data in results:
        for col_num, value in enumerate(voter_data, 1):
            ws.cell(row=row, column=col_num, value=value)
        row += 1

        if row % 5000 == 0:
            print(f"    ...{row-6:,} unmatched voters written")

    auto_adjust_columns(ws)
    print(f"  OK Unmatched Voters: {len(results):,} voters")

def get_output_dir(district_type, district_number):
    """Get the output directory for a specific district, creating subfolder if needed"""
    district_folder = BASE_OUTPUT_DIR / f"{district_type}_{district_number}"
    district_folder.mkdir(exist_ok=True, parents=True)
    return district_folder

def clean_old_files(district_type, district_number, keep_latest=3):
    """Delete old Excel files for this district, keeping only the most recent ones"""

    output_dir = get_output_dir(district_type, district_number)

    # Find all files for this district
    pattern = f"{district_type}_{district_number}_Complete_Analysis_*.xlsx"
    matching_files = sorted(output_dir.glob(pattern), key=lambda x: x.stat().st_mtime, reverse=True)

    if len(matching_files) <= keep_latest:
        print(f"Found {len(matching_files)} existing file(s) for {district_type} {district_number} (keeping all)")
        return

    # Files to delete (keep the newest keep_latest files)
    files_to_delete = matching_files[keep_latest:]

    print(f"\nCleaning up old files for {district_type} {district_number}:")
    print(f"  Found: {len(matching_files)} files")
    print(f"  Keeping: {keep_latest} newest files")
    print(f"  Deleting: {len(files_to_delete)} old files\n")

    for old_file in files_to_delete:
        try:
            old_file.unlink()
            print(f"  Deleted: {old_file.name}")
        except Exception as e:
            print(f"  [!] Could not delete {old_file.name}: {e}")


def create_modeled_ethnicity_tab(wb, conn, district_type, district_number):

    """
    Ethnicity breakdown using ModeledEthnicity (100% coverage).
    Shows: total voters, matched vs unmatched counts, and party split per ethnicity.
    """
    col = _COL_MAP[district_type]

    ws = wb.create_sheet("Ethnicity (Modeled)")
    ws.sheet_properties.tabColor = "7030A0"

    # â"€â"€ Title â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    _lbl = 'Statewide' if district_type == 'STATEWIDE' else f'{district_type} {district_number}'
    ws["A1"] = f"{_lbl} \u2013 Ethnicity Breakdown (ModeledEthnicity)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")
    ws["A2"] = f"Source: ModeledEthnicity column • Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A2:I2")

    # â"€â"€ Section 1: Matched vs Unmatched by Ethnicity â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    ws["A4"] = "SECTION 1 – Audience Match Rate by Ethnicity"
    ws["A4"].font = Font(bold=True, size=12, color="1F497D")
    ws.merge_cells("A4:G4")

    headers1 = ["Ethnicity", "Total Voters", "% of District",
                "In Audiences", "Match Rate", "Not in Audiences", "Miss Rate"]
    for ci, h in enumerate(headers1, 1):
        ws.cell(row=5, column=ci, value=h)
    format_header_row(ws, 5)

    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT
                COALESCE(NULLIF(ModeledEthnicity,''), 'Unknown') AS eth,
                COUNT(*) AS total,
                SUM(CASE WHEN origin IS NOT NULL AND TRIM(origin)!='' THEN 1 ELSE 0 END) AS matched,
                SUM(CASE WHEN origin IS NULL OR TRIM(origin)=''  THEN 1 ELSE 0 END) AS unmatched
            FROM voter_file
            WHERE {_dc(col)}
            GROUP BY 1 ORDER BY 2 DESC
        """, _dp(col, district_number))
        rows = cur.fetchall()

    grand_total = sum(int(r[1]) for r in rows)

    # Ethnicity color map
    eth_colors = {
        "White / Caucasian":       "DDEEFF",
        "Hispanic / Latino":       "FFF2CC",
        "Asian / Pacific Islander":"E2EFDA",
        "Black / African American":"FCE4D6",
        "Other / Multi-Racial":    "EDE7F6",
        "Unknown":                 "F2F2F2",
    }

    row = 6
    for r in rows:
        eth = str(r[0]); total = int(r[1]); matched = int(r[2]); unmatched = int(r[3])
        pct_district = total * 100.0 / grand_total if grand_total else 0
        match_rate   = matched * 100.0 / total if total else 0
        miss_rate    = unmatched * 100.0 / total if total else 0
        bg = eth_colors.get(eth, "FFFFFF")
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        vals = [eth, total, round(pct_district, 1), matched,
                round(match_rate, 1), unmatched, round(miss_rate, 1)]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.fill = fill
            if ci in (2, 4, 6):
                cell.number_format = "#,##0"
            if ci in (3, 5, 7):
                cell.number_format = "0.0"
        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=grand_total).number_format = "#,##0"
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 2

    # â"€â"€ Section 2: Party Registration by Ethnicity â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    ws.cell(row=row, column=1, value="SECTION 2 – Party Registration by Ethnicity")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F497D")
    ws.merge_cells(f"A{row}:I{row}")
    row += 1

    # Get all party values present
    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT DISTINCT OfficialParty FROM voter_file
            WHERE {_dc(col)} AND OfficialParty IS NOT NULL
            ORDER BY OfficialParty
        """, _dp(col, district_number))
        all_parties = [r[0] for r in cur.fetchall()]

    # Consolidate minor parties into "Other"
    major = {"Democrat", "Republican", "Unaffiliated", "Conservative", "Working Families"}
    party_cols = [p for p in all_parties if p in major] + ["Other"]

    # Header row
    hdr_row = row
    ws.cell(row=hdr_row, column=1, value="Ethnicity")
    for ci, p in enumerate(party_cols, 2):
        ws.cell(row=hdr_row, column=ci, value=p)
    last_col = len(party_cols) + 2
    ws.cell(row=hdr_row, column=last_col, value="Total")
    format_header_row(ws, hdr_row)
    row += 1

    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT
                COALESCE(NULLIF(ModeledEthnicity,''), 'Unknown') AS eth,
                OfficialParty,
                COUNT(*) AS n
            FROM voter_file
            WHERE {_dc(col)}
            GROUP BY 1, 2
        """, _dp(col, district_number))
        party_raw = cur.fetchall()

    from collections import defaultdict
    party_data = defaultdict(lambda: defaultdict(int))
    eth_totals  = defaultdict(int)
    for eth, party, n in party_raw:
        eth = str(eth) if eth else "Unknown"
        n = int(n)
        bucket = party if party in major else "Other"
        party_data[eth][bucket] += n
        eth_totals[eth] += n

    eth_order = [str(r[0]) for r in rows]  # same order as section 1
    for eth in eth_order:
        bg = eth_colors.get(eth, "FFFFFF")
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        ws.cell(row=row, column=1, value=eth).fill = fill
        for ci, p in enumerate(party_cols, 2):
            cell = ws.cell(row=row, column=ci, value=party_data[eth].get(p, 0))
            cell.number_format = "#,##0"
            cell.fill = fill
        ws.cell(row=row, column=last_col, value=eth_totals[eth]).number_format = "#,##0"
        ws.cell(row=row, column=last_col).fill = fill
        row += 1

    auto_adjust_columns(ws)
    print(f"  OK Ethnicity (Modeled) tab created – {grand_total:,} voters, {len(rows)} groups")


def create_boe_donor_tab(wb, conn, district_type, district_number):
    """BOE Donors tab: per-year D/R/U breakdown, donor summary, full donor list."""
    import datetime as _dt
    col = _COL_MAP[district_type]
    ws = wb.create_sheet("BOE Donors")
    ws.sheet_properties.tabColor = "375623"

    YEAR_MAX = _dt.date.today().year
    YEAR_MIN = YEAR_MAX - 9
    YEARS    = list(range(YEAR_MIN, YEAR_MAX + 1))

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    sec_font = Font(bold=True, size=12, color="1F497D")
    sec_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    D_fill   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    R_fill   = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
    U_fill   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    tot_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    num_fmt  = "#,##0"
    amt_fmt  = "$#,##0.00"
    pct_fmt  = "0.0%"

    ws["A1"] = f"{'Statewide' if district_type == 'STATEWIDE' else f'{district_type} {district_number}'} - BOE Donor Analysis (State Campaign Finance)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:N1")
    ws["A2"] = f"Source: NYS Board of Elections | Years: {YEAR_MIN}-{YEAR_MAX} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A2:N2")

    with conn.cursor() as cur:
        cur.execute(f"SELECT COUNT(*) FROM voter_file WHERE {_dc(col)}", _dp(col, district_number))
        total_voters = cur.fetchone()[0]

    # ----------------------------------------------------------------
    # SECTION 1: Summary by Party
    # ----------------------------------------------------------------
    row = 4
    ws.cell(row=row, column=1, value="SECTION 1 - DONOR SUMMARY BY PARTY").font = sec_font
    for ci in range(1, 7): ws.cell(row=row, column=ci).fill = sec_fill
    row += 1

    sum_hdrs = ["Party", "Donors", "% of District", "Total $", "Avg $", "Contributions"]
    for ci, h in enumerate(sum_hdrs, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    row += 1

    with conn.cursor() as cur:
        grand_donors = grand_amt = grand_cnt = 0
        for label, amt_col, cnt_col, fill in [
            ("Democratic",   "boe_total_D_amt", "boe_total_D_count", D_fill),
            ("Republican",   "boe_total_R_amt", "boe_total_R_count", R_fill),
            ("Unaffiliated", "boe_total_U_amt", "boe_total_U_count", U_fill),
        ]:
            cur.execute(f"""
                SELECT COUNT(*), COALESCE(SUM({amt_col}),0), COALESCE(SUM({cnt_col}),0)
                FROM voter_file WHERE {_dc(col)} AND {amt_col} IS NOT NULL AND {amt_col} > 0
            """, _dp(col, district_number))
            donors, amt, cnt = cur.fetchone()
            donors = int(donors or 0); amt = float(amt or 0); cnt = int(cnt or 0)
            grand_donors += donors; grand_amt += amt; grand_cnt += cnt
            vals = [label, donors, donors/total_voters if total_voters else 0,
                    amt, amt/donors if donors else 0, cnt]
            for ci2, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci2, value=v)
                cell.fill = fill
                if ci2 == 2: cell.number_format = num_fmt
                if ci2 == 3: cell.number_format = pct_fmt
                if ci2 in (4, 5): cell.number_format = amt_fmt
                if ci2 == 6: cell.number_format = num_fmt
            row += 1

        # Grand total row
        cur.execute(f"""
            SELECT COUNT(*), COALESCE(SUM(boe_total_amt),0), COALESCE(SUM(boe_total_count),0)
            FROM voter_file WHERE {_dc(col)} AND boe_total_amt IS NOT NULL AND boe_total_amt > 0
        """, _dp(col, district_number))
        td, ta, tc = cur.fetchone()
        td = int(td or 0); ta = float(ta or 0); tc = int(tc or 0)
        vals = ["TOTAL", td, td/total_voters if total_voters else 0,
                ta, ta/td if td else 0, tc]
        for ci2, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci2, value=v)
            cell.fill = tot_fill
            cell.font = Font(bold=True)
            if ci2 == 2: cell.number_format = num_fmt
            if ci2 == 3: cell.number_format = pct_fmt
            if ci2 in (4, 5): cell.number_format = amt_fmt
            if ci2 == 6: cell.number_format = num_fmt
        row += 2

    # ----------------------------------------------------------------
    # SECTION 2: Year-by-Year Breakdown
    # ----------------------------------------------------------------
    ws.cell(row=row, column=1, value="SECTION 2 - YEAR-BY-YEAR BREAKDOWN (matched donors in this district)").font = sec_font
    for ci in range(1, 10): ws.cell(row=row, column=ci).fill = sec_fill
    row += 1

    yr_hdrs = ["Year", "Dem $", "Dem #", "Rep $", "Rep #", "Unaf $", "Unaf #", "Total $", "Total #"]
    for ci, h in enumerate(yr_hdrs, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    row += 1

    # Pull per-year totals by joining voter_file -> boe_donor_summary for this district
    with conn.cursor() as cur:
        yr_totals = {}
        for yr in YEARS:
            cur.execute(f"""
                SELECT
                    COALESCE(SUM(b.y{yr}_D_amt),   0), COALESCE(SUM(b.y{yr}_D_count), 0),
                    COALESCE(SUM(b.y{yr}_R_amt),   0), COALESCE(SUM(b.y{yr}_R_count), 0),
                    COALESCE(SUM(b.y{yr}_U_amt),   0), COALESCE(SUM(b.y{yr}_U_count), 0)
                FROM nys_voter_tagging.voter_file v
                JOIN boe_donors.boe_donor_summary b ON v.StateVoterId = b.StateVoterId
                WHERE {_dc(col)}
            """, _dp(col, district_number))
            da, dc, ra, rc, ua, uc = cur.fetchone()
            yr_totals[yr] = (float(da),int(dc), float(ra),int(rc), float(ua),int(uc))

    grand_yr_d_amt = grand_yr_d_cnt = 0.0
    grand_yr_r_amt = grand_yr_r_cnt = 0.0
    grand_yr_u_amt = grand_yr_u_cnt = 0.0

    for yr in YEARS:
        da,dc,ra,rc,ua,uc = yr_totals[yr]
        ta = da+ra+ua; tc2 = dc+rc+uc
        grand_yr_d_amt += da; grand_yr_d_cnt += dc
        grand_yr_r_amt += ra; grand_yr_r_cnt += rc
        grand_yr_u_amt += ua; grand_yr_u_cnt += uc
        vals = [yr, da, dc, ra, rc, ua, uc, ta, tc2]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            if ci in (2, 4, 6, 8): cell.number_format = amt_fmt
            if ci in (3, 5, 7, 9): cell.number_format = num_fmt
        row += 1

    # Year grand total
    gt_vals = ["TOTAL",
               grand_yr_d_amt, int(grand_yr_d_cnt),
               grand_yr_r_amt, int(grand_yr_r_cnt),
               grand_yr_u_amt, int(grand_yr_u_cnt),
               grand_yr_d_amt+grand_yr_r_amt+grand_yr_u_amt,
               int(grand_yr_d_cnt+grand_yr_r_cnt+grand_yr_u_cnt)]
    for ci, v in enumerate(gt_vals, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.fill = tot_fill; cell.font = Font(bold=True)
        if ci in (2, 4, 6, 8): cell.number_format = amt_fmt
        if ci in (3, 5, 7, 9): cell.number_format = num_fmt
    row += 2

    # ----------------------------------------------------------------
    # SECTION 3: Donor List
    # ----------------------------------------------------------------

    has_email = _has_crm_email(conn)
    NUM_COLS = 15 + (1 if has_email else 0)

    ws.cell(row=row, column=1, value="SECTION 3 - DONOR LIST").font = sec_font
    for ci in range(1, NUM_COLS + 1): ws.cell(row=row, column=ci).fill = sec_fill
    row += 1

    list_hdrs = ["StateVoterId", "FirstName", "LastName", "Address", "City", "ZIP",
                 "Phone", "Landline", "Mobile", "Reg Party",
                 "Dem Total", "Rep Total", "Unaf Total", "Last Date", "Last Filer"]
    if has_email:
        list_hdrs.append("Email")
    for ci, h in enumerate(list_hdrs, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    row += 1

    D_row_fill   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    R_row_fill   = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
    CON_row_fill = PatternFill(start_color="E2CFED", end_color="E2CFED", fill_type="solid")

    D_hdr_fill   = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    R_hdr_fill   = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    CON_hdr_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    OTH_hdr_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    sec_hdr_font = Font(bold=True, size=11, color="FFFFFF")
    sub_fill     = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sub_font     = Font(bold=True, italic=True)

    email_col_bare = ", crm_email" if has_email else ""

    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT StateVoterId, FirstName, LastName,
                   PrimaryAddress1, PrimaryCity, PrimaryZip,
                   PrimaryPhone, Landline, Mobile,
                   OfficialParty,
                   COALESCE(boe_total_D_amt, 0),
                   COALESCE(boe_total_R_amt, 0),
                   COALESCE(boe_total_U_amt, 0),
                   boe_last_date,
                   boe_last_filer{email_col_bare}
            FROM voter_file
            WHERE {_dc(col)}
              AND boe_total_amt IS NOT NULL AND boe_total_amt > 0
            ORDER BY OfficialParty, boe_total_amt DESC
            LIMIT 200000
        """, _dp(col, district_number))
        donors = cur.fetchall()

    PARTY_ORDER = ["Republican", "Conservative", "Democrat"]
    def party_sort_key(rec):
        p = rec[9] or ""
        try: return (PARTY_ORDER.index(p), 0)
        except ValueError: return (len(PARTY_ORDER), 0)
    donors = sorted(donors, key=party_sort_key)

    from itertools import groupby
    for party, group_iter in groupby(donors, key=lambda r: r[9] or "Other"):
        group_rows = list(group_iter)
        if party == "Democrat":       hfill = D_hdr_fill;   row_fill = D_row_fill
        elif party == "Republican":   hfill = R_hdr_fill;   row_fill = R_row_fill
        elif party == "Conservative": hfill = CON_hdr_fill; row_fill = CON_row_fill
        else:                         hfill = OTH_hdr_fill; row_fill = None

        lbl = ws.cell(row=row, column=1, value=f"--- {party.upper()} ---")
        lbl.font = sec_hdr_font; lbl.fill = hfill
        for ci in range(2, NUM_COLS + 1): ws.cell(row=row, column=ci).fill = hfill
        row += 1

        grp_d = grp_r = grp_u = 0.0
        for r_data in group_rows:
            for ci, v in enumerate(r_data, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                if row_fill: cell.fill = row_fill
                if ci in (11, 12, 13): cell.number_format = amt_fmt
            if party == "Republican":
                for ci in range(1, NUM_COLS + 1):
                    ws.cell(row=row, column=ci).font = Font(bold=True, size=11)
            # Highlight Democrats who donated to Republicans (yellow/gold)
            if party == "Democrat" and float(r_data[11] or 0) > 0:
                _xover_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                for ci in range(1, NUM_COLS + 1):
                    ws.cell(row=row, column=ci).fill = _xover_fill
                    ws.cell(row=row, column=ci).font = Font(bold=True, size=11)
            grp_d += float(r_data[10] or 0)
            grp_r += float(r_data[11] or 0)
            grp_u += float(r_data[12] or 0)
            row += 1

        ws.cell(row=row, column=1, value=f"  {party} SUBTOTAL ({len(group_rows):,} donors)").font = sub_font
        ws.cell(row=row, column=1).fill = sub_fill
        for ci, v, fmt in [(11, grp_d, amt_fmt), (12, grp_r, amt_fmt), (13, grp_u, amt_fmt)]:
            cell = ws.cell(row=row, column=ci, value=v)
            cell.number_format = fmt; cell.font = sub_font; cell.fill = sub_fill
        other_cols = [2,3,4,5,6,7,8,9,10,14,15] + ([NUM_COLS] if has_email else [])
        for ci in other_cols: ws.cell(row=row, column=ci).fill = sub_fill
        row += 2

    auto_adjust_columns(ws)
    print(f"  OK BOE Donors tab: {len(donors):,} donors, {len(YEARS)}-year breakdown")

def create_national_donor_tab(wb, conn, district_type, district_number):
    """National donor tab with detailed party contribution breakdown."""
    col = _COL_MAP[district_type]
    ws = wb.create_sheet("National Donor")
    ws.sheet_properties.tabColor = "203864"

    has_email = _has_crm_email(conn)
    email_col_bare = ", crm_email" if has_email else ""
    NUM_COLS_NAT = 19 if has_email else 18

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    sec_font = Font(bold=True, size=12, color="1F497D")
    sec_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    num_fmt  = "#,##0"
    amt_fmt  = "$#,##0.00"
    pct_fmt  = "0.0%"

    ws["A1"] = f"{'Statewide' if district_type == 'STATEWIDE' else f'{district_type} {district_number}'} - National Donor Analysis"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:M1")
    ws["A2"] = f"Source: National Contributions (2020-2024) | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A2:M2")

    row = 4
    ws.cell(row=row, column=1, value="SUMMARY - Contributions by Party Signal").font = sec_font
    for ci in range(1, 9): ws.cell(row=row, column=ci).fill = sec_fill
    row += 1

    # Enhanced summary with party breakdown
    sum_hdrs = ["Party Signal", "Donors", "% of District", "Total $", "Avg $", "Contributions", "Avg/Donor"]
    for ci, h in enumerate(sum_hdrs, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    row += 1

    with conn.cursor() as cur:
        cur.execute(f"SELECT COUNT(*) FROM voter_file WHERE {_dc(col)}", _dp(col, district_number))
        total_voters = cur.fetchone()[0]

        # Get totals for each party signal
        for party_label, amt_col, cnt_col in [
            ("Democratic", "national_democratic_amount", "national_democratic_count"),
            ("Republican", "national_republican_amount", "national_republican_count"),
            ("Independent", "national_independent_amount", "national_independent_count"),
            ("Unknown", "national_unknown_amount", "national_unknown_count")
        ]:
            cur.execute(f"""
                SELECT
                    COUNT(DISTINCT StateVoterId),
                    COALESCE(SUM({amt_col}), 0),
                    COALESCE(AVG({amt_col}), 0),
                    COALESCE(SUM({cnt_col}), 0)
                FROM voter_file
                WHERE {_dc(col)} AND {cnt_col} > 0
            """, _dp(col, district_number))
            r = cur.fetchone()
            donors = int(r[0] or 0)
            total = float(r[1] or 0)
            avg_amt = float(r[2] or 0)
            contributions = int(r[3] or 0)
            
            vals = [
                party_label,
                donors,
                donors/total_voters if total_voters else 0,
                total,
                avg_amt,
                contributions,
                contributions/donors if donors else 0
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                if ci == 2: cell.number_format = num_fmt
                if ci == 3: cell.number_format = pct_fmt
                if ci in (4, 5): cell.number_format = amt_fmt
                if ci in (6, 7): cell.number_format = num_fmt
            row += 1
        
        # Overall total row
        cur.execute(f"""
            SELECT 
                COUNT(DISTINCT StateVoterId),
                COALESCE(SUM(national_total_amount), 0),
                COALESCE(AVG(national_total_amount), 0),
                COALESCE(SUM(national_total_count), 0)
            FROM voter_file
            WHERE {_dc(col)} AND is_national_donor = TRUE
        """, (*_dp(col, district_number),))
        r = cur.fetchone()
        all_donors = int(r[0] or 0)
        all_total = float(r[1] or 0)
        all_avg = float(r[2] or 0)
        all_contribs = int(r[3] or 0)
        
        ws.cell(row=row, column=1, value="TOTAL (all parties)").font = Font(bold=True)
        vals = [
            all_donors,
            all_donors/total_voters if total_voters else 0,
            all_total,
            all_avg,
            all_contribs,
            all_contribs/all_donors if all_donors else 0
        ]
        for ci, v in enumerate(vals, 2):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = Font(bold=True)
            if ci == 2: cell.number_format = num_fmt
            if ci == 3: cell.number_format = pct_fmt
            if ci in (4, 5): cell.number_format = amt_fmt
            if ci in (6, 7): cell.number_format = num_fmt
    
    row += 2

    # Donor list with party breakdown columns
    ws.cell(row=row, column=1, value="DONOR LIST - Detailed Contribution Breakdown").font = sec_font
    for ci in range(1, NUM_COLS_NAT + 1): ws.cell(row=row, column=ci).fill = sec_fill
    row += 1

    list_hdrs = ["StateVoterId", "FirstName", "LastName", "Address", "City", "ZIP",
                 "Phone", "Reg Party",
                 "Dem $", "Dem #", "Rep $", "Rep #", "Ind $", "Ind #",
                 "Total $", "Total #", "LD", "SD"]
    if has_email:
        list_hdrs.append("Email")
    for ci, h in enumerate(list_hdrs, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    row += 1

    D_row_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    R_row_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")

    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT StateVoterId, FirstName, LastName,
                   PrimaryAddress1, PrimaryCity, PrimaryZip,
                   PrimaryPhone, OfficialParty,
                   COALESCE(national_democratic_amount, 0), COALESCE(national_democratic_count, 0),
                   COALESCE(national_republican_amount, 0), COALESCE(national_republican_count, 0),
                   COALESCE(national_independent_amount, 0), COALESCE(national_independent_count, 0),
                   COALESCE(national_total_amount, 0), COALESCE(national_total_count, 0),
                   LDName, SDName{email_col_bare}
            FROM voter_file
            WHERE {_dc(col)} AND is_national_donor = TRUE
            ORDER BY OfficialParty, national_total_amount DESC
            LIMIT 200000
        """, (*_dp(col, district_number),))
        donors = cur.fetchall()

    PARTY_ORDER = ["Republican", "Conservative", "Democrat"]
    def party_sort_key(rec):
        p = rec[7] or ""  # OfficialParty is now at index 7
        try: return (PARTY_ORDER.index(p), 0)
        except ValueError: return (len(PARTY_ORDER), 0)
    donors = sorted(donors, key=party_sort_key)

    sec_hdr_font = Font(bold=True, size=11, color="FFFFFF")
    D_hdr_fill   = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    R_hdr_fill   = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    CON_hdr_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    OTH_hdr_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    CON_row_fill = PatternFill(start_color="E2CFED", end_color="E2CFED", fill_type="solid")

    # Group donors by party with subtotals
    from itertools import groupby
    for party, group_iter in groupby(donors, key=lambda r: r[7] or 'Other'):
        group_rows = list(group_iter)
        if party == 'Democrat':       hfill = D_hdr_fill;   row_fill = D_row_fill
        elif party == 'Republican':   hfill = R_hdr_fill;   row_fill = R_row_fill
        elif party == 'Conservative': hfill = CON_hdr_fill; row_fill = CON_row_fill
        else:                         hfill = OTH_hdr_fill; row_fill = None

        # Party header row
        lbl = ws.cell(row=row, column=1, value=f'--- {party.upper()} ---')
        lbl.font = sec_hdr_font; lbl.fill = hfill
        for ci in range(2, NUM_COLS_NAT + 1): ws.cell(row=row, column=ci).fill = hfill
        row += 1

        grp_dem_amt = grp_rep_amt = grp_ind_amt = grp_total_amt = 0.0
        grp_dem_cnt = grp_rep_cnt = grp_ind_cnt = grp_total_cnt = 0
        
        for r_data in group_rows:
            for ci, v in enumerate(r_data, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                if row_fill: cell.fill = row_fill
                if ci in (9, 11, 13, 15): cell.number_format = amt_fmt  # Amount columns
                if ci in (10, 12, 14, 16): cell.number_format = num_fmt  # Count columns
            
            if party == 'Republican':
                for ci in range(1, NUM_COLS_NAT + 1):
                    ws.cell(row=row, column=ci).font = Font(bold=True, size=11)
            # Highlight Democrats who donated to Republicans (yellow/gold)
            if party == 'Democrat' and float(r_data[10] or 0) > 0:
                _xover_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                for ci in range(1, NUM_COLS_NAT + 1):
                    ws.cell(row=row, column=ci).fill = _xover_fill
                    ws.cell(row=row, column=ci).font = Font(bold=True, size=11)

            grp_dem_amt += float(r_data[8] or 0)
            grp_dem_cnt += int(r_data[9] or 0)
            grp_rep_amt += float(r_data[10] or 0)
            grp_rep_cnt += int(r_data[11] or 0)
            grp_ind_amt += float(r_data[12] or 0)
            grp_ind_cnt += int(r_data[13] or 0)
            grp_total_amt += float(r_data[14] or 0)
            grp_total_cnt += int(r_data[15] or 0)
            row += 1

        # Subtotal row
        sub_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        sub_font = Font(bold=True, italic=True)
        ws.cell(row=row, column=1, value=f'  {party} SUBTOTAL ({len(group_rows):,} donors)').font = sub_font
        ws.cell(row=row, column=1).fill = sub_fill
        
        # Subtotal values
        subtotal_vals = [
            (9, grp_dem_amt, amt_fmt), (10, grp_dem_cnt, num_fmt),
            (11, grp_rep_amt, amt_fmt), (12, grp_rep_cnt, num_fmt),
            (13, grp_ind_amt, amt_fmt), (14, grp_ind_cnt, num_fmt),
            (15, grp_total_amt, amt_fmt), (16, grp_total_cnt, num_fmt)
        ]
        for col_idx, val, fmt in subtotal_vals:
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.number_format = fmt
            cell.font = sub_font
            cell.fill = sub_fill
        
        # Fill other cells in subtotal row
        for ci in [2,3,4,5,6,7,8,17,18] + ([19] if has_email else []):
            ws.cell(row=row, column=ci).fill = sub_fill
        row += 2

    auto_adjust_columns(ws)
    print(f"  OK National tab: {len(donors):,} National donors with party breakdown")


def create_party_voter_tab(wb, conn, district_type, district_number,
                           tab_name, party_codes, header_color):
    """Create a tab listing all voter file info for voters of specified party registration(s)."""
    col = _COL_MAP[district_type]
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = header_color

    has_email = _has_crm_email(conn)
    has_phone = _has_crm_phone(conn)

    # Title
    ws['A1'] = f"{'Statewide' if district_type == 'STATEWIDE' else f'{district_type} {district_number}'} - {tab_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:L1')
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'] = "Total Voters:"
    ws['A3'].font = Font(bold=True)

    # Build headers
    headers = [
        'StateVoterId', 'FirstName', 'MiddleName', 'LastName', 'Suffix',
        'Address', 'City', 'State', 'ZIP',
        'DOB', 'Party',
        'Phone', 'Landline', 'Mobile',
        'County', 'LD', 'SD', 'CD',
        'Reg Date', 'Last Activity',
        'Ethnicity',
    ]
    if has_email:
        headers.append('CRM Email')
    if has_phone:
        headers.extend(['CRM Phone', 'CRM Mobile'])

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Build query
    placeholders = ','.join(['%s'] * len(party_codes))
    crm_cols = ""
    if has_email:
        crm_cols += ", crm_email"
    if has_phone:
        crm_cols += ", crm_phone, crm_mobile"
    query = f"""
        SELECT StateVoterId, FirstName, MiddleName, LastName, SuffixName,
               PrimaryAddress1, PrimaryCity, PrimaryState, PrimaryZip,
               DOB, OfficialParty,
               PrimaryPhone, Landline, Mobile,
               CountyName, LDName, SDName, CDName,
               RegistrationDate, LastVoterActivity,
               ModeledEthnicity
               {crm_cols}
        FROM voter_file
        WHERE {_dc(col)} AND OfficialParty IN ({placeholders})
        ORDER BY LastName, FirstName
        LIMIT 200000
    """

    with conn.cursor() as cur:
        cur.execute(query, (*_dp(col, district_number), *party_codes))
        results = cur.fetchall()

    ws['B3'] = len(results)
    ws['B3'].number_format = '#,##0'

    if not results:
        ws['A7'] = "No voters found for this party in this district"
        auto_adjust_columns(ws)
        print(f"  OK {tab_name}: 0 voters")
        return

    row = 6
    for record in results:
        for ci, value in enumerate(record, 1):
            ws.cell(row=row, column=ci, value=value)
        row += 1
        if (row - 6) % 10000 == 0:
            print(f"    ...{row-6:,} voters written")

    auto_adjust_columns(ws)
    print(f"  OK {tab_name}: {len(results):,} voters")


def create_cfb_donor_tab(wb, conn, district_type, district_number):
    """NYC CFB (Campaign Finance Board) donor tab with per-cycle breakdown."""
    col = _COL_MAP[district_type]
    ws = wb.create_sheet("CFB Donors")
    ws.sheet_properties.tabColor = "4A235A"

    has_email = _has_crm_email(conn)
    email_col_bare = ", crm_email" if has_email else ""
    NUM_COLS_CFB = 17 if has_email else 16

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="4A235A", end_color="4A235A", fill_type="solid")
    sec_font = Font(bold=True, size=12, color="4A235A")
    sec_fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
    tot_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    num_fmt  = "#,##0"
    amt_fmt  = "$#,##0.00"
    pct_fmt  = "0.0%"

    ws["A1"] = f"{'Statewide' if district_type == 'STATEWIDE' else f'{district_type} {district_number}'} - NYC CFB Donor Analysis (City Campaign Finance)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:L1")
    ws["A2"] = f"Source: NYC Campaign Finance Board | Cycles: 2017, 2021, 2023, 2025 | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A2:L2")

    with conn.cursor() as cur:
        cur.execute(f"SELECT COUNT(*) FROM voter_file WHERE {_dc(col)}", (*_dp(col, district_number),))
        total_voters = cur.fetchone()[0]

    # ----------------------------------------------------------------
    # SECTION 1: Summary
    # ----------------------------------------------------------------
    row = 4
    ws.cell(row=row, column=1, value="SECTION 1 - CFB DONOR SUMMARY").font = sec_font
    for ci in range(1, 7): ws.cell(row=row, column=ci).fill = sec_fill
    row += 1

    sum_hdrs = ["Metric", "Value"]
    for ci, h in enumerate(sum_hdrs, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    row += 1

    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT COUNT(*),
                   COALESCE(SUM(cfb_total_amt), 0),
                   COALESCE(AVG(cfb_total_amt), 0),
                   COALESCE(SUM(cfb_total_count), 0)
            FROM voter_file
            WHERE {_dc(col)} AND cfb_total_amt IS NOT NULL AND cfb_total_amt > 0
        """, (*_dp(col, district_number),))
        donors, total_amt, avg_amt, total_count = cur.fetchone()
        donors = int(donors or 0)
        total_amt = float(total_amt or 0)
        avg_amt = float(avg_amt or 0)
        total_count = int(total_count or 0)

    metrics = [
        ("Total CFB Donors", donors, num_fmt),
        ("% of " + ("State" if district_type == "STATEWIDE" else "District"), donors / total_voters if total_voters else 0, pct_fmt),
        ("Total Donated", total_amt, amt_fmt),
        ("Average Donation", avg_amt, amt_fmt),
        ("Total Contributions", total_count, num_fmt),
    ]
    for label, val, fmt in metrics:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = fmt
        row += 1
    row += 1

    # ----------------------------------------------------------------
    # SECTION 2: Per-Cycle Breakdown
    # ----------------------------------------------------------------
    ws.cell(row=row, column=1, value="SECTION 2 - CYCLE BREAKDOWN").font = sec_font
    for ci in range(1, 5): ws.cell(row=row, column=ci).fill = sec_fill
    row += 1

    cycle_hdrs = ["Cycle", "Donors", "Total $", "Avg $"]
    for ci, h in enumerate(cycle_hdrs, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    row += 1

    CYCLES = [2017, 2021, 2023, 2025]
    grand_donors = grand_amt = 0
    with conn.cursor() as cur:
        for yr in CYCLES:
            amt_col = f"cfb_{yr}_amt"
            cur.execute(f"""
                SELECT COUNT(*), COALESCE(SUM({amt_col}), 0), COALESCE(AVG({amt_col}), 0)
                FROM voter_file
                WHERE {_dc(col)} AND {amt_col} IS NOT NULL AND {amt_col} > 0
            """, (*_dp(col, district_number),))
            d, a, av = cur.fetchone()
            d = int(d or 0); a = float(a or 0); av = float(av or 0)
            grand_donors += d; grand_amt += a
            vals = [yr, d, a, av]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                if ci == 2: cell.number_format = num_fmt
                if ci in (3, 4): cell.number_format = amt_fmt
            row += 1

    # Total row
    vals = ["TOTAL", grand_donors, grand_amt, grand_amt / grand_donors if grand_donors else 0]
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.fill = tot_fill; cell.font = Font(bold=True)
        if ci == 2: cell.number_format = num_fmt
        if ci in (3, 4): cell.number_format = amt_fmt
    row += 2

    # ----------------------------------------------------------------
    # SECTION 3: Donor List
    # ----------------------------------------------------------------
    ws.cell(row=row, column=1, value="SECTION 3 - CFB DONOR LIST").font = sec_font
    for ci in range(1, NUM_COLS_CFB + 1): ws.cell(row=row, column=ci).fill = sec_fill
    row += 1

    list_hdrs = ["StateVoterId", "FirstName", "LastName", "Address", "City", "ZIP",
                 "Phone", "Reg Party", "CFB Total $", "CFB Count",
                 "2017 $", "2021 $", "2023 $", "2025 $",
                 "Last Candidate", "Last Office"]
    if has_email:
        list_hdrs.append("Email")
    for ci, h in enumerate(list_hdrs, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    row += 1

    D_row_fill   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    R_row_fill   = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
    CON_row_fill = PatternFill(start_color="E2CFED", end_color="E2CFED", fill_type="solid")
    D_hdr_fill   = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    R_hdr_fill   = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    CON_hdr_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    OTH_hdr_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    sec_hdr_font = Font(bold=True, size=11, color="FFFFFF")

    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT StateVoterId, FirstName, LastName,
                   PrimaryAddress1, PrimaryCity, PrimaryZip,
                   PrimaryPhone, OfficialParty,
                   COALESCE(cfb_total_amt, 0), COALESCE(cfb_total_count, 0),
                   COALESCE(cfb_2017_amt, 0), COALESCE(cfb_2021_amt, 0),
                   COALESCE(cfb_2023_amt, 0), COALESCE(cfb_2025_amt, 0),
                   cfb_last_cand, cfb_last_office{email_col_bare}
            FROM voter_file
            WHERE {_dc(col)} AND cfb_total_amt IS NOT NULL AND cfb_total_amt > 0
            ORDER BY OfficialParty, cfb_total_amt DESC
            LIMIT 200000
        """, (*_dp(col, district_number),))
        donors = cur.fetchall()

    PARTY_ORDER = ["Republican", "Conservative", "Democrat"]
    def party_sort_key(rec):
        p = rec[7] or ""
        try: return (PARTY_ORDER.index(p), 0)
        except ValueError: return (len(PARTY_ORDER), 0)
    donors = sorted(donors, key=party_sort_key)

    from itertools import groupby
    for party, group_iter in groupby(donors, key=lambda r: r[7] or "Other"):
        group_rows = list(group_iter)
        if party == "Democrat":       hfill = D_hdr_fill;   row_fill = D_row_fill
        elif party == "Republican":   hfill = R_hdr_fill;   row_fill = R_row_fill
        elif party == "Conservative": hfill = CON_hdr_fill; row_fill = CON_row_fill
        else:                         hfill = OTH_hdr_fill; row_fill = None

        lbl = ws.cell(row=row, column=1, value=f"--- {party.upper()} ---")
        lbl.font = sec_hdr_font; lbl.fill = hfill
        for ci in range(2, NUM_COLS_CFB + 1): ws.cell(row=row, column=ci).fill = hfill
        row += 1

        grp_total = 0.0
        grp_count = 0
        for r_data in group_rows:
            for ci, v in enumerate(r_data, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                if row_fill: cell.fill = row_fill
                if ci in (9, 11, 12, 13, 14): cell.number_format = amt_fmt
                if ci == 10: cell.number_format = num_fmt
            if party == "Republican":
                for ci in range(1, NUM_COLS_CFB + 1):
                    ws.cell(row=row, column=ci).font = Font(bold=True, size=11)
            grp_total += float(r_data[8] or 0)
            grp_count += int(r_data[9] or 0)
            row += 1

        sub_fill2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        sub_font = Font(bold=True, italic=True)
        ws.cell(row=row, column=1, value=f"  {party} SUBTOTAL ({len(group_rows):,} donors)").font = sub_font
        ws.cell(row=row, column=1).fill = sub_fill2
        cell = ws.cell(row=row, column=9, value=grp_total)
        cell.number_format = amt_fmt; cell.font = sub_font; cell.fill = sub_fill2
        cell = ws.cell(row=row, column=10, value=grp_count)
        cell.number_format = num_fmt; cell.font = sub_font; cell.fill = sub_fill2
        for ci in [2,3,4,5,6,7,8,11,12,13,14,15,16] + ([17] if has_email else []):
            ws.cell(row=row, column=ci).fill = sub_fill2
        row += 2

    auto_adjust_columns(ws)
    print(f"  OK CFB Donors tab: {len(donors):,} NYC CFB donors with cycle breakdown")


def create_explanation_tab(wb, district_type, district_number, has_cfb=False):
    """First tab: explains every sheet in the workbook and notes on turnout overlap."""
    ws = wb.create_sheet("Guide", 0)
    ws.sheet_properties.tabColor = "000000"

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color="1F497D")
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    note_font = Font(italic=True, size=11, color="C00000")
    hdr_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")
    xover_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 90

    ws["A1"] = f"{'Statewide' if district_type == 'STATEWIDE' else f'{district_type} {district_number}'} - Export Guide"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")
    from datetime import datetime as _dt
    ws["A2"] = f"Generated: {_dt.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A2:B2")

    row = 4
    ws.cell(row=row, column=1, value="SHEET DESCRIPTIONS").font = section_font
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    for ci, h in enumerate(["Sheet Name", "Description"], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    row += 1

    sheets = [
        ("Guide", "This sheet. Describes every tab in this workbook."),
        ("Summary", "Overview of all audience files matched to this district. Shows unique voter counts per audience, "
                     "combined deduplicated totals (so double-counted voters are removed), and overall district match rate."),
        ("Ethnicity (Modeled)", "Demographic breakdown using surname-based ethnicity modeling. "
                                "Section 1 shows audience match rates by ethnicity. "
                                "Section 2 shows party registration breakdown by ethnicity."),
        ("BOE Donors", "NYS Board of Elections state-level campaign finance donors matched to voters in this district. "
                       "Section 1: summary by party. Section 2: year-by-year totals. "
                       "Section 3: full donor list grouped by party registration (Republicans first). "
                       "Gold highlighted rows = registered Democrats who also donated to Republican candidates."),
        ("National Donor", "National campaign contributions matched to voters. "
                           "Summary section shows totals by party signal (Dem/Rep/Ind/Unknown). "
                           "Donor list grouped by voter registration party (Republicans first). "
                           "Gold highlighted rows = registered Democrats who also donated to Republican candidates/committees."),
    ]
    if has_cfb:
        sheets.append(
            ("CFB Donors", "NYC Campaign Finance Board city-level contributions. "
                           "Shows per-cycle breakdown (2017, 2021, 2023, 2025) and full donor list "
                           "grouped by voter registration party (Republicans first).")
        )
    sheets += [
        ("Registered Democrats", "Full voter roster of all registered Democrats in this district "
                                 "with contact info, address, DOB, registration date, ethnicity, and email (if available)."),
        ("Republicans & Conservatives", "Full voter roster of all registered Republicans and Conservatives in this district "
                                        "with contact info, address, DOB, registration date, ethnicity, and email (if available)."),
        ("Issue Audience Tabs (orange)", "One tab per issue-based audience file (e.g. 'blue collar', 'border security crisis', "
                                         "'children in home'). Each tab lists every voter in this district who matched that audience. "
                                         "A voter can appear on MULTIPLE issue tabs if they matched multiple audiences."),
        ("Turnout Model Tabs (green)", "One tab per turnout model (HT = High Turnout, MT = Medium Turnout, LT = Low Turnout "
                                       "for DEM/GOP/IND). Each tab lists voters scored into that turnout tier."),
        ("Unmatched Voters", "Voters in this district who did not match ANY audience file. "
                             "These are voters with no audience data from the Causeway match."),
    ]

    for i, (name, desc) in enumerate(sheets):
        fill = alt_fill if i % 2 == 0 else None
        c1 = ws.cell(row=row, column=1, value=name)
        c1.font = bold_font
        c2 = ws.cell(row=row, column=2, value=desc)
        c2.font = normal_font
        c2.alignment = Alignment(wrap_text=True)
        if fill:
            c1.fill = fill; c2.fill = fill
        row += 1

    row += 1

    # --- COLOR KEY section ---
    ws.cell(row=row, column=1, value="COLOR KEY (Donor Tabs)").font = section_font
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    colors = [
        (PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid"),
         "Republican (voter registration)"),
        (PatternFill(start_color="E2CFED", end_color="E2CFED", fill_type="solid"),
         "Conservative (voter registration)"),
        (PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
         "Democrat (voter registration)"),
        (xover_fill,
         "CROSSOVER: Registered Democrat who donated to Republican candidates (gold highlight, bold text)"),
    ]
    for fill, label in colors:
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).value = ""
        ws.cell(row=row, column=2, value=label).font = normal_font
        row += 1

    row += 1

    # --- TURNOUT OVERLAP NOTE ---
    ws.cell(row=row, column=1, value="IMPORTANT: TURNOUT AUDIENCE OVERLAP").font = section_font
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    overlap_notes = [
        "Turnout models (HT/MT/LT) are scored independently for each party propensity (DEM, GOP, IND).",
        "A single voter CAN appear on multiple turnout tabs. For example, a voter might be scored as "
        "HT HARD GOP (high-turnout strong Republican) AND also appear on MT SOFT DEM (medium-turnout weak Democrat) "
        "if their voting history and demographics produce overlapping signals.",
        "This is by design: the models measure LIKELIHOOD of supporting each party at different turnout levels, "
        "not a single mutually exclusive assignment.",
        "The 'COMBINED UNIQUE' row on the Summary tab shows the deduplicated count across all turnout models, "
        "removing any double-counting.",
        "Issue audiences can also overlap with each other and with turnout models. "
        "A voter in 'border security crisis' can simultaneously be in 'blue collar' and 'HT HARD GOP'.",
        "When planning outreach, use the COMBINED UNIQUE counts for accurate universe sizing, "
        "not the sum of individual audience counts.",
    ]
    for note in overlap_notes:
        ws.cell(row=row, column=1, value="").font = normal_font
        c = ws.cell(row=row, column=2, value=note)
        c.font = normal_font
        c.alignment = Alignment(wrap_text=True)
        row += 1

    print("  OK Guide tab created")


def main():
    parser = argparse.ArgumentParser(description='Export district analysis to Excel')
    parser.add_argument('--ld', type=str, help='Legislative District number (e.g., 63 or 063)')
    parser.add_argument('--sd', type=str, help='State Senate District number (e.g., 5 or 05)')
    parser.add_argument('--cd', type=str, help='Congressional District number (e.g., 3 or 03)')
    parser.add_argument('--county', type=str, help='County name (e.g., Albany, Erie, Kings)')
    parser.add_argument('--statewide', action='store_true', help='Export all voters with no geographic filter')
    parser.add_argument('--keep', type=int, default=0, help='Number of recent files to keep (default: 0 - delete all old files)')
    parser.add_argument('--no-clean', action='store_true', help='Skip cleaning old files')
    args = parser.parse_args()

    # Determine which district type was specified
    district_type = None
    district_number = None
    district_number_padded = None  # For filename

    if args.statewide:
        district_type = 'STATEWIDE'
        district_number = None
        district_number_padded = 'Statewide'
    elif args.ld:
        district_type = 'LD'
        district_number_padded = args.ld
        district_number = str(int(args.ld))  # Remove leading zeros for DB query
    elif args.sd:
        district_type = 'SD'
        district_number_padded = args.sd
        district_number = str(int(args.sd))  # Remove leading zeros for DB query
    elif args.cd:
        district_type = 'CD'
        district_number_padded = args.cd
        district_number = str(int(args.cd))  # Remove leading zeros for DB query
    elif args.county:
        district_type = 'COUNTY'
        district_number_padded = args.county.title()  # Capitalize for consistency (e.g., Albany, Erie)
        district_number = args.county.title()  # County names don't have leading zeros
    else:
        # Default to LD 63
        district_type = 'LD'
        district_number = '63'
        district_number_padded = '063'
        print("No district specified, defaulting to LD 63")

    label = 'STATEWIDE (all voters)' if district_type == 'STATEWIDE' else f'{district_type} {district_number}'
    print(f"\n{'='*80}")
    print(f"  EXPORTING {label} TO EXCEL - ONE SHEET PER AUDIENCE FILE")
    if district_type != 'STATEWIDE':
        print(f"  Database query: {district_type}Name = '{district_number}'")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*80}\n")

    # Clean old files unless --no-clean is specified
    if not args.no_clean:
        clean_old_files(district_type, district_number_padded, keep_latest=args.keep)

    conn = connect_db()

    try:
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create summary tab
        print("Creating Summary tab...")
        audience_list = create_summary_tab(wb, conn, district_type, district_number)

        # Create ethnicity tab (modeled)
        print("Creating Ethnicity (Modeled) tab...")
        create_modeled_ethnicity_tab(wb, conn, district_type, district_number)

        # Check if BOE donor columns exist - REQUIRED
        has_boe_donors = False
        try:
            with conn.cursor() as cur:
                # Check for all required BOE columns
                required_cols = ['boe_total_D_amt', 'boe_total_R_amt', 'boe_total_U_amt']
                cur.execute(f"SELECT {', '.join(required_cols)} FROM voter_file LIMIT 0")
                has_boe_donors = True
        except:
            pass
        
        if not has_boe_donors:
            print("\n" + "="*80)
            print("  ERROR: DONOR DATA MISSING")
            print("="*80)
            print("\nBOE donor columns not found in voter_file.")
            print("\nYou must run donor enrichment before exporting:")
            print("\n  python main.py donors")
            print("\nThis will add both BOE (state) and National donor data.")
            print("="*80 + "\n")
            sys.exit(1)

        # Pre-compute CRM email on voter_file (cached; instant after first run)
        from pipeline.crm_merge import enrich_voter_email
        enrich_voter_email(conn)

        print('Creating BOE Donors tab...')
        create_boe_donor_tab(wb, conn, district_type, district_number)

        # Check if National donor columns exist - REQUIRED
        has_national_donors = False
        try:
            with conn.cursor() as cur:
                # Check for all required National columns
                required_cols = ['is_national_donor', 'national_total_amount', 'national_democratic_amount', 'national_republican_amount']
                cur.execute(f"SELECT {', '.join(required_cols)} FROM voter_file LIMIT 0")
                has_national_donors = True
        except:
            pass

        if not has_national_donors:
            print("\n" + "="*80)
            print("  ERROR: NATIONAL DONOR DATA MISSING")
            print("="*80)
            print("\nNational donor columns not found in voter_file.")
            print("\nYou must run donor enrichment before exporting:")
            print("\n  python main.py donors")
            print("\nThis will add both BOE (state) and National donor data.")
            print("="*80 + "\n")
            sys.exit(1)
        
        print('Creating National Donor tab...')
        create_national_donor_tab(wb, conn, district_type, district_number)

        # Check if CFB donor columns exist (optional - NYC districts only)
        has_cfb_donors = False
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT cfb_total_amt FROM voter_file LIMIT 0")
                has_cfb_donors = True
        except Exception:
            pass

        if has_cfb_donors:
            print('Creating CFB Donors tab...')
            create_cfb_donor_tab(wb, conn, district_type, district_number)
        else:
            print('  Skipping CFB Donors tab (cfb columns not found - run: python main.py cfb-enrich)')

        # Create explanation / guide tab (inserted at position 0, before Summary)
        print("Creating Guide tab...")
        create_explanation_tab(wb, district_type, district_number, has_cfb=has_cfb_donors)
        # Move Guide to be the very first sheet
        wb.move_sheet("Guide", offset=-(len(wb.sheetnames)-1))


        # Party voter roster tabs
        print('Creating Registered Democrats tab...')
        create_party_voter_tab(wb, conn, district_type, district_number,
                               "Registered Democrats", ["Democrat"], "1F4E79")

        print('Creating Republicans & Conservatives tab...')
        create_party_voter_tab(wb, conn, district_type, district_number,
                               "Republicans & Conservatives", ["Republican", "Conservative"], "8B0000")

        # Split audience list into issues and turnout, create in order
        _TP = ('HT ', 'MT ', 'LT ')
        issues_list  = [(a, v) for a, v in audience_list if not a.upper().startswith(_TP)]
        turnout_list = [(a, v) for a, v in audience_list if a.upper().startswith(_TP)]

        print(f"\nCreating {len(issues_list)} issue tabs...")
        for i, (audience_file, voters) in enumerate(issues_list, 1):
            tab_name = audience_file.replace('.csv', '').replace('INDV NYS_', '')[:31]
            print(f"\n[{i}/{len(issues_list)}] {audience_file}")
            create_audience_tab(wb, conn, district_type, district_number, audience_file, tab_name, tab_color="E26B0A")

        print(f"\nCreating {len(turnout_list)} turnout tabs...")
        for i, (audience_file, voters) in enumerate(turnout_list, 1):
            tab_name = audience_file.replace('.csv', '').replace('INDV NYS_', '')[:31]
            print(f"\n[{i}/{len(turnout_list)}] {audience_file}")
            create_audience_tab(wb, conn, district_type, district_number, audience_file, tab_name, tab_color="00B050")

        # Create unmatched voters tab
        print("\nCreating Unmatched Voters tab...")
        create_unmatched_tab(wb, conn, district_type, district_number)

        # Get output directory and save workbook (use padded number for filename)
        output_dir = get_output_dir(district_type, district_number_padded)
        output_file = output_dir / f"{district_type}_{district_number_padded}_Complete_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print(f"\nSaving workbook...")
        wb.save(output_file)

        print(f"\n{'='*80}")
        print(f"  SUCCESS!")
        print(f"{'='*80}")
        print(f"\nFile saved to:")
        print(f"  {output_file}")

        print(f"\nWorkbook contains:")
        print("  - Ethnicity (Modeled) tab")
        print("  - BOE Donors tab")
        print("  - National Donor tab")
        if has_cfb_donors:
            print("  - CFB Donors tab")
        print("  - Registered Democrats tab")
        print("  - Republicans & Conservatives tab")
        print(f"  - {len(audience_list)} unique audience file tabs")
        print(f"  - Unmatched Voters tab")
        print(f"\nTotal tabs: {len(wb.worksheets)}")
        print(f"\nNOTE: Voters appearing in multiple audiences will be listed on multiple sheets")
        print()

    finally:
        conn.close()

if __name__ == "__main__":
    main()
