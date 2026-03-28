#!/usr/bin/env python3
"""
Export Voter Contact workbook — a variant of the full district export with:
- No turnout model tabs (HT/MT/LT audiences)
- Party sheets for Republicans, Democrats, Other (all columns)
- Contact-method tabs: Mobile / Landline / Email, each split by party
- CRM-enriched email, phone, and mobile where available

Usage:
    python export_contact.py --ld 63
    python export_contact.py --sd 5
    python export_contact.py --cd 3
    python export_contact.py --county Albany

Called by: python main.py voter-contact --ld 63
"""

import os
import sys
from pathlib import Path
from datetime import datetime
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))

# Reuse shared functions from the main export module
from export.export import (
    connect_db,
    format_header_row,
    auto_adjust_columns,
    _has_crm_email,
    _has_crm_phone,
    create_boe_donor_tab,
    create_national_donor_tab,
    create_cfb_donor_tab,
    get_output_dir,
    BASE_OUTPUT_DIR,
)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
# Columns to exclude from the "all columns" party sheets
_EXCLUDE_COLS = {"origin", "clean_first", "clean_last",
                  "StateEthnicity", "ModeledEthnicity", "ObservedEthnicity"}

# Friendly header overrides for raw DB column names
_HEADER_MAP = {
    "StateVoterId": "State Voter ID",
    "FirstName": "First Name",
    "MiddleName": "Middle Name",
    "LastName": "Last Name",
    "SuffixName": "Suffix",
    "PrimaryAddress1": "Address",
    "PrimaryCity": "City",
    "PrimaryState": "State",
    "PrimaryZip": "ZIP",
    "PrimaryZip4": "ZIP4",
    "PrimaryOddEvenCode": "Odd/Even",
    "PrimaryHouseNumber": "House #",
    "PrimaryHouseHalf": "House Half",
    "PrimaryStreetPre": "Street Pre",
    "PrimaryStreetName": "Street Name",
    "PrimaryStreetType": "Street Type",
    "PrimaryStreetPost": "Street Post",
    "PrimaryUnit": "Unit",
    "PrimaryUnitNumber": "Unit #",
    "SecondaryAddress1": "Secondary Address",
    "SecondaryCity": "Secondary City",
    "SecondaryState": "Secondary State",
    "SecondaryZip": "Secondary ZIP",
    "SecondaryZip4": "Secondary ZIP4",
    "SecondaryUnit": "Secondary Unit",
    "SecondaryUnitNumber": "Secondary Unit #",
    "PrimaryPhone": "Phone",
    "PrimaryPhoneTRC": "Phone TRC",
    "UserLandline": "User Landline",
    "LandlineTRC": "Landline TRC",
    "LandlineDNC": "Landline DNC",
    "HasPrimaryPhone": "Has Phone",
    "MobileTRC": "Mobile TRC",
    "UserMobile": "User Mobile",
    "MobileDNC": "Mobile DNC",
    "OfficialParty": "Party",
    "CalculatedParty": "Calculated Party",
    "HouseholdParty": "Household Party",
    "ObservedParty": "Observed Party",
    "RegistrationDate": "Reg Date",
    "RegistrationStatus": "Reg Status",
    "LastVoterActivity": "Last Activity",
    "GeneralFrequency": "Gen Freq",
    "PrimaryFrequency": "Pri Freq",
    "OverAllFrequency": "Overall Freq",
    "GeneralRegularity": "Gen Regularity",
    "PrimaryRegularity": "Pri Regularity",
    "GeneralAbsenteeStatus": "Gen Absentee",
    "PrimaryAbsenteeStatus": "Pri Absentee",
    "AbsenteeStatus": "Absentee Status",
    "MailSortCodeRoute": "Mail Route",
    "MailDeliveryPt": "Mail Delivery Pt",
    "MailDeliveryPtChkDigit": "Mail Chk Digit",
    "MailLineOfTravel": "Mail Line Travel",
    "MailLineOfTravelOrder": "Mail Travel Order",
    "MailDPVStatus": "Mail DPV Status",
    "NeighborhoodId": "Neighborhood ID",
    "NeighborhoodSegmentId": "Neighborhood Seg",
    "CountyName": "County",
    "CountyNumber": "County #",
    "PrecinctNumber": "Precinct #",
    "PrecinctName": "Precinct",
    "CDName": "CD",
    "LDName": "LD",
    "SDName": "SD",
    "CensusBlock": "Census Block",
    "JurisdictionalVoterId": "Jurisdictional ID",
    "RNCRegId": "RNC Reg ID",
    "StateEthnicity": "State Ethnicity",
    "ModeledEthnicity": "Modeled Ethnicity",
    "ObservedEthnicity": "Observed Ethnicity",
    "AgeRange": "Age Range",
    "HHRecId": "HH Rec ID",
    "HHMemberId": "HH Member ID",
    "HHCode": "HH Code",
    "VoterKey": "Voter Key",
    "MapCode": "Map Code",
    "ClientId": "Client ID",
    # Enriched columns
    "crm_email": "CRM Email",
    "crm_phone": "CRM Phone",
    "crm_mobile": "CRM Mobile",
    "boe_total_amt": "BOE Total $",
    "boe_total_count": "BOE # Donations",
    "boe_total_D_amt": "BOE Dem $",
    "boe_total_R_amt": "BOE Rep $",
    "boe_total_U_amt": "BOE Other $",
    "boe_last_date": "BOE Last Date",
    "boe_last_filer": "BOE Last Filer",
    "is_national_donor": "National Donor?",
    "national_total_amount": "National Total $",
    "national_democratic_amount": "National Dem $",
    "national_republican_amount": "National Rep $",
    "national_last_date": "National Last Date",
    "cfb_total_amt": "CFB Total $",
    "cfb_total_count": "CFB # Donations",
    "cfb_last_date": "CFB Last Date",
    "cfb_last_cand": "CFB Last Candidate",
}

# Party definitions — used for both party tabs and contact-method tabs
_REP_SQL = "OfficialParty = %s"
_REP_PARAMS = ("Republican",)
_DEM_SQL = "OfficialParty = %s"
_DEM_PARAMS = ("Democrat",)
_OTH_SQL = "OfficialParty NOT IN (%s, %s)"
_OTH_PARAMS = ("Republican", "Democrat")

# Contact-method WHERE fragments (built dynamically based on available columns)
def _mobile_filter(has_crm):
    """SQL fragment for voters with any mobile number."""
    parts = ["(UserMobile IS NOT NULL AND UserMobile != '')"]
    if has_crm:
        parts.append("(crm_mobile IS NOT NULL AND crm_mobile != '')")
    return "(" + " OR ".join(parts) + ")"


def _landline_filter(has_crm):
    """SQL fragment for voters with any phone/landline number."""
    parts = [
        "(PrimaryPhone IS NOT NULL AND PrimaryPhone != '')",
        "(UserLandline IS NOT NULL AND UserLandline != '')",
    ]
    if has_crm:
        parts.append("(crm_phone IS NOT NULL AND crm_phone != '')")
    return "(" + " OR ".join(parts) + ")"


def _email_filter(has_crm):
    """SQL fragment for voters with an email address."""
    if has_crm:
        return "(crm_email IS NOT NULL AND crm_email != '')"
    # No CRM columns — no email data exists
    return "1=0"


# ---------------------------------------------------------------------------
# Guide tab
# ---------------------------------------------------------------------------
def create_contact_guide_tab(wb, district_type, district_number, has_cfb=False):
    """Guide tab tailored for the Voter Contact workbook."""
    ws = wb.create_sheet("Guide", 0)
    ws.sheet_properties.tabColor = "000000"

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color="1F497D")
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    hdr_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")
    xover_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 90

    ws["A1"] = f"{district_type} {district_number} - Voter Contact Guide"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A2:B2")

    row = 4
    ws.cell(row=row, column=1, value="SHEET DESCRIPTIONS").font = section_font
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    for ci, h in enumerate(["Sheet Name", "Description"], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    row += 1

    sheets = [
        ("Guide", "This sheet. Describes every tab in this Voter Contact workbook."),
        ("Summary", "District overview: party registration breakdown, contact method coverage "
                     "(mobile, landline, email, CRM), and donor source counts with percentages."),
        ("BOE Donors", "NYS Board of Elections state-level campaign finance donors matched to voters in this district. "
                       "Summary by party, year-by-year totals, and full donor list grouped by party. "
                       "Gold highlighted rows = registered Democrats who also donated to Republican candidates."),
        ("National Donor", "National campaign contributions matched to voters. "
                           "Summary shows totals by party signal. Donor list grouped by voter registration party. "
                           "Gold highlighted rows = Dem-registered voters who donated to Republican candidates/committees."),
    ]
    if has_cfb:
        sheets.append(
            ("CFB Donors", "NYC Campaign Finance Board city-level contributions. "
                           "Per-cycle breakdown and full donor list grouped by voter registration party.")
        )
    sheets += [
        # --- Party tabs ---
        ("Registered Republicans", "Full voter roster of all registered Republicans in this district with ALL voter_file "
                                   "columns including CRM-enriched email, phone, and mobile where available."),
        ("Registered Democrats", "Full voter roster of all registered Democrats in this district with ALL voter_file "
                                 "columns including CRM-enriched email, phone, and mobile where available."),
        ("Other Parties", "Full voter roster of all other registered voters (Conservative, Working Families, Independence, "
                          "Libertarian, Green, Blank, etc.) with ALL voter_file columns and CRM contact enrichment."),
        # --- Contact method tabs ---
        ("Mobile - Republicans", "Registered Republicans who have a mobile number (UserMobile or CRM Mobile). All columns."),
        ("Mobile - Democrats", "Registered Democrats who have a mobile number (UserMobile or CRM Mobile). All columns."),
        ("Mobile - Other", "Other-party voters who have a mobile number (UserMobile or CRM Mobile). All columns."),
        ("Landline - Republicans", "Registered Republicans who have a phone number (PrimaryPhone, UserLandline, or CRM Phone). All columns."),
        ("Landline - Democrats", "Registered Democrats who have a phone number (PrimaryPhone, UserLandline, or CRM Phone). All columns."),
        ("Landline - Other", "Other-party voters who have a phone number (PrimaryPhone, UserLandline, or CRM Phone). All columns."),
        ("Email - Republicans", "Registered Republicans who have a CRM email address. All columns."),
        ("Email - Democrats", "Registered Democrats who have a CRM email address. All columns."),
        ("Email - Other", "Other-party voters who have a CRM email address. All columns."),
    ]

    for i, (name, desc) in enumerate(sheets):
        fill = alt_fill if i % 2 == 0 else None
        c1 = ws.cell(row=row, column=1, value=name)
        c1.font = bold_font
        c2 = ws.cell(row=row, column=2, value=desc)
        c2.font = normal_font
        c2.alignment = Alignment(wrap_text=True)
        if fill:
            c1.fill = fill
            c2.fill = fill
        row += 1

    row += 1

    # --- COLOR KEY ---
    ws.cell(row=row, column=1, value="COLOR KEY").font = section_font
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    colors = [
        (PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid"),
         "Republican tabs (dark red) — party roster + contact-method tabs"),
        (PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
         "Democrat tabs (navy blue) — party roster + contact-method tabs"),
        (PatternFill(start_color="7B5EA7", end_color="7B5EA7", fill_type="solid"),
         "Other Parties tabs (purple) — party roster + contact-method tabs"),
        (xover_fill,
         "CROSSOVER: Registered Democrat who donated to Republican candidates (gold, bold, donor tabs only)"),
    ]
    for fill_c, label in colors:
        ws.cell(row=row, column=1).fill = fill_c
        ws.cell(row=row, column=1).value = ""
        ws.cell(row=row, column=2, value=label).font = normal_font
        row += 1

    row += 1

    # --- CRM ENRICHMENT NOTE ---
    ws.cell(row=row, column=1, value="CRM CONTACT ENRICHMENT").font = section_font
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    crm_notes = [
        "Party and contact-method tabs include CRM Email, CRM Phone, and CRM Mobile columns from the crm_unified database.",
        "CRM data is matched to voter records by name (first + last) against HubSpot and Campaign Monitor contacts.",
        "Not every voter will have CRM data. Empty cells mean no CRM match was found for that voter.",
        "Contact-method tabs (Mobile/Landline/Email) filter to voters who have at least one value in that channel, "
        "from EITHER the voter file source columns OR CRM enrichment.",
        "To improve CRM coverage, ensure contacts in HubSpot / Campaign Monitor have accurate first and last names.",
    ]
    for note in crm_notes:
        ws.cell(row=row, column=2, value=note).font = normal_font
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        row += 1

    print("  OK Voter Contact Guide tab created")


# ---------------------------------------------------------------------------
# Summary tab — district overview with party, contact, and donor stats
# ---------------------------------------------------------------------------
def create_contact_summary_tab(wb, conn, district_type, district_number):
    """Summary tab with district stats: party breakdown, contact coverage, donors."""
    dist_col = {"LD": "LDName", "SD": "SDName", "CD": "CDName", "COUNTY": "CountyName"}[district_type]
    ws = wb.create_sheet("Summary", 1)  # insert right after Guide
    ws.sheet_properties.tabColor = "366092"

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color="1F497D")
    hdr_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    num_fmt = "#,##0"
    pct_fmt = "0.0%"
    R_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
    D_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    O_fill = PatternFill(start_color="E2BCFA", end_color="E2BCFA", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14

    ws["A1"] = f"{district_type} {district_number} - Summary"
    ws["A1"].font = title_font
    ws.merge_cells("A1:C1")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A2:C2")

    base_where = f"`{dist_col}` = %s"
    base_params = (district_number,)

    with conn.cursor() as cur:
        cur.execute(f"SELECT COUNT(*) FROM voter_file WHERE {base_where}", base_params)
        total_voters = cur.fetchone()[0]

        # ── Section 1: Party Registration ─────────────────────────────────
        row = 4
        ws.cell(row=row, column=1, value="PARTY REGISTRATION").font = section_font
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        for ci, h in enumerate(["Party", "Voters", "% of District"], 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
        row += 1

        parties = [
            ("Republican", _REP_SQL, _REP_PARAMS, R_fill),
            ("Democrat",   _DEM_SQL, _DEM_PARAMS, D_fill),
            ("Other",      _OTH_SQL, _OTH_PARAMS, O_fill),
        ]
        for label, psql, pparams, fill in parties:
            cur.execute(f"SELECT COUNT(*) FROM voter_file WHERE {base_where} AND {psql}",
                        base_params + pparams)
            cnt = cur.fetchone()[0]
            ws.cell(row=row, column=1, value=label).font = bold_font
            c2 = ws.cell(row=row, column=2, value=cnt)
            c2.number_format = num_fmt
            c3 = ws.cell(row=row, column=3, value=cnt / total_voters if total_voters else 0)
            c3.number_format = pct_fmt
            for ci in range(1, 4):
                ws.cell(row=row, column=ci).fill = fill
            row += 1

        # Total row
        ws.cell(row=row, column=1, value="Total Registered").font = bold_font
        c2 = ws.cell(row=row, column=2, value=total_voters)
        c2.font = bold_font; c2.number_format = num_fmt
        row += 2

        # ── Section 2: Contact Coverage ───────────────────────────────────
        ws.cell(row=row, column=1, value="CONTACT COVERAGE").font = section_font
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        for ci, h in enumerate(["Contact Method", "Voters", "% of District"], 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
        row += 1

        has_crm = _has_crm_email(conn)
        contact_methods = [
            ("Mobile", _mobile_filter(has_crm)),
            ("Landline", _landline_filter(has_crm)),
            ("Email", _email_filter(has_crm)),
        ]
        for i, (label, sql_frag) in enumerate(contact_methods):
            cur.execute(f"SELECT COUNT(*) FROM voter_file WHERE {base_where} AND {sql_frag}",
                        base_params)
            cnt = cur.fetchone()[0]
            fill = alt_fill if i % 2 == 0 else None
            ws.cell(row=row, column=1, value=label).font = bold_font
            c2 = ws.cell(row=row, column=2, value=cnt)
            c2.number_format = num_fmt
            c3 = ws.cell(row=row, column=3, value=cnt / total_voters if total_voters else 0)
            c3.number_format = pct_fmt
            if fill:
                for ci in range(1, 4):
                    ws.cell(row=row, column=ci).fill = fill
            row += 1

        # CRM enriched count
        if has_crm:
            cur.execute(f"SELECT COUNT(*) FROM voter_file WHERE {base_where} "
                        "AND crm_email IS NOT NULL", base_params)
            crm_cnt = cur.fetchone()[0]
            ws.cell(row=row, column=1, value="CRM Matched").font = bold_font
            c2 = ws.cell(row=row, column=2, value=crm_cnt)
            c2.number_format = num_fmt
            c3 = ws.cell(row=row, column=3, value=crm_cnt / total_voters if total_voters else 0)
            c3.number_format = pct_fmt
            row += 1

        row += 1

        # ── Section 3: Donor Coverage ─────────────────────────────────────
        ws.cell(row=row, column=1, value="DONOR COVERAGE").font = section_font
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        for ci, h in enumerate(["Source", "Donors", "% of District"], 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
        row += 1

        donor_sources = [
            ("BOE (State)", "boe_total_amt IS NOT NULL AND boe_total_amt > 0"),
            ("National (FEC)", "is_national_donor = 1"),
        ]
        # CFB only if column exists
        try:
            cur.execute("SELECT cfb_total_amt FROM voter_file LIMIT 0")
            donor_sources.append(("CFB (NYC)", "cfb_total_amt IS NOT NULL AND cfb_total_amt > 0"))
        except Exception:
            pass

        for i, (label, donor_sql) in enumerate(donor_sources):
            cur.execute(f"SELECT COUNT(*) FROM voter_file WHERE {base_where} AND {donor_sql}",
                        base_params)
            cnt = cur.fetchone()[0]
            fill = alt_fill if i % 2 == 0 else None
            ws.cell(row=row, column=1, value=label).font = bold_font
            c2 = ws.cell(row=row, column=2, value=cnt)
            c2.number_format = num_fmt
            c3 = ws.cell(row=row, column=3, value=cnt / total_voters if total_voters else 0)
            c3.number_format = pct_fmt
            if fill:
                for ci in range(1, 4):
                    ws.cell(row=row, column=ci).fill = fill
            row += 1

    auto_adjust_columns(ws)
    print("  OK Summary tab created")


# ---------------------------------------------------------------------------
# Column discovery
# ---------------------------------------------------------------------------
def _discover_voter_columns(conn):
    """Return the ordered list of voter_file columns, excluding internal ones."""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT COLUMN_NAME FROM information_schema.COLUMNS "
            "WHERE TABLE_SCHEMA = 'nys_voter_tagging' AND TABLE_NAME = 'voter_file' "
            "ORDER BY ORDINAL_POSITION"
        )
        return [r[0] for r in cur.fetchall() if r[0] not in _EXCLUDE_COLS]


# ---------------------------------------------------------------------------
# Full-column party / contact tab
# ---------------------------------------------------------------------------
def create_full_party_tab(wb, conn, district_type, district_number,
                          tab_name, party_sql, party_params, header_color,
                          columns, extra_sql=None):
    """Create a tab with ALL voter_file columns for voters matching filters.

    Parameters
    ----------
    extra_sql : str, optional
        Additional SQL AND fragment (e.g. contact-method filter).
        Injected as raw SQL — must be safe (built internally, never from user input).
    """
    dist_col = {"LD": "LDName", "SD": "SDName", "CD": "CDName", "COUNTY": "CountyName"}[district_type]
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = header_color

    # Title rows
    ws['A1'] = f"{district_type} {district_number} - {tab_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:L1')
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'] = "Total Voters:"
    ws['A3'].font = Font(bold=True)

    # Headers (row 5)
    headers = [_HEADER_MAP.get(c, c) for c in columns]
    fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    # Build query
    col_list = ", ".join(f"`{c}`" for c in columns)
    where = f"`{dist_col}` = %s AND {party_sql}"
    if extra_sql:
        where += f" AND {extra_sql}"
    query = f"SELECT {col_list} FROM voter_file WHERE {where} ORDER BY LastName, FirstName"
    params = (district_number,) + party_params

    with conn.cursor() as cur:
        cur.execute(query, params)
        results = cur.fetchall()

    ws['B3'] = len(results)
    ws['B3'].number_format = '#,##0'

    if not results:
        ws['A7'] = "No voters found for this filter in this district"
        auto_adjust_columns(ws)
        print(f"  OK {tab_name}: 0 voters")
        return

    row = 6
    for record in results:
        for ci, value in enumerate(record, 1):
            ws.cell(row=row, column=ci, value=value)
        row += 1
        if (row - 6) % 25000 == 0:
            print(f"    ...{row-6:,} voters written")

    auto_adjust_columns(ws)
    print(f"  OK {tab_name}: {len(results):,} voters")


# ---------------------------------------------------------------------------
# File cleanup
# ---------------------------------------------------------------------------
def clean_old_contact_files(district_type, district_number, keep_latest=3):
    """Delete old Voter Contact Excel files for this district."""
    output_dir = get_output_dir(district_type, district_number)
    pattern = f"{district_type}_{district_number}_Voter_Contact_*.xlsx"
    matching = sorted(output_dir.glob(pattern), key=lambda x: x.stat().st_mtime, reverse=True)
    if len(matching) <= keep_latest:
        print(f"Found {len(matching)} existing Voter Contact file(s) (keeping all)")
        return
    to_delete = matching[keep_latest:]
    print(f"\nCleaning up old Voter Contact files:")
    print(f"  Keeping: {keep_latest} newest, Deleting: {len(to_delete)} old\n")
    for f in to_delete:
        try:
            f.unlink()
            print(f"  Deleted: {f.name}")
        except Exception as e:
            print(f"  [!] Could not delete {f.name}: {e}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Export Voter Contact workbook")
    parser.add_argument("--ld", type=str, help="Legislative District number")
    parser.add_argument("--sd", type=str, help="State Senate District number")
    parser.add_argument("--cd", type=str, help="Congressional District number")
    parser.add_argument("--county", type=str, help="County name")
    parser.add_argument("--keep", type=int, default=0, help="Number of recent files to keep")
    parser.add_argument("--no-clean", action="store_true", help="Skip cleaning old files")
    parser.add_argument("--quiet", action="store_true")
    parser.add_argument("--verbose", action="store_true")
    parser.add_argument("--debug", action="store_true")
    args = parser.parse_args()

    # Determine district
    district_type = district_number = district_number_padded = None
    if args.ld:
        district_type = "LD"
        district_number_padded = args.ld
        district_number = str(int(args.ld))
    elif args.sd:
        district_type = "SD"
        district_number_padded = args.sd
        district_number = str(int(args.sd))
    elif args.cd:
        district_type = "CD"
        district_number_padded = args.cd
        district_number = str(int(args.cd))
    elif args.county:
        district_type = "COUNTY"
        district_number_padded = args.county.title()
        district_number = args.county.title()
    else:
        district_type = "LD"
        district_number = "63"
        district_number_padded = "063"
        print("No district specified, defaulting to LD 63")

    print(f"\n{'='*80}")
    print(f"  VOTER CONTACT EXPORT — {district_type} {district_number}")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*80}\n")

    if not args.no_clean:
        clean_old_contact_files(district_type, district_number_padded, keep_latest=args.keep)

    conn = connect_db()

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ── Donor tabs ─────────────────────────────────────────────────────
        # BOE donors (required)
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT boe_total_D_amt, boe_total_R_amt, boe_total_U_amt FROM voter_file LIMIT 0")
        except Exception:
            print("\nERROR: BOE donor columns not found. Run: python main.py donors")
            conn.close()
            sys.exit(1)

        # CRM enrichment (email + phone + mobile)
        from pipeline.crm_merge import enrich_voter_crm
        enrich_voter_crm(conn)

        print("Creating BOE Donors tab...")
        create_boe_donor_tab(wb, conn, district_type, district_number)

        # National donors (required)
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT is_national_donor, national_total_amount FROM voter_file LIMIT 0")
        except Exception:
            print("\nERROR: National donor columns not found. Run: python main.py donors")
            conn.close()
            sys.exit(1)

        print("Creating National Donor tab...")
        create_national_donor_tab(wb, conn, district_type, district_number)

        # CFB donors (optional)
        has_cfb = False
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT cfb_total_amt FROM voter_file LIMIT 0")
                has_cfb = True
        except Exception:
            pass

        if has_cfb:
            print("Creating CFB Donors tab...")
            create_cfb_donor_tab(wb, conn, district_type, district_number)
        else:
            print("  Skipping CFB Donors tab (columns not found)")

        # Guide tab (created at position 0)
        print("Creating Voter Contact Guide tab...")
        create_contact_guide_tab(wb, district_type, district_number, has_cfb=has_cfb)

        # ── Discover columns ─────────────────────────────────────────────────
        print("\nDiscovering voter_file columns...")
        columns = _discover_voter_columns(conn)
        print(f"  {len(columns)} columns available\n")

        # Detect CRM column availability for contact-method filters
        has_crm = _has_crm_email(conn)

        # ── Party tabs: Rep, Dem, Other ──────────────────────────────────────
        party_tabs = [
            ("Registered Republicans", _REP_SQL, _REP_PARAMS, "8B0000"),
            ("Registered Democrats",   _DEM_SQL, _DEM_PARAMS, "1F4E79"),
            ("Other Parties",          _OTH_SQL, _OTH_PARAMS, "7B5EA7"),
        ]
        for tab_name, party_sql, party_params, color in party_tabs:
            print(f"Creating {tab_name} tab...")
            create_full_party_tab(
                wb, conn, district_type, district_number,
                tab_name, party_sql, party_params, color, columns,
            )

        # ── Contact-method tabs: Mobile / Landline / Email × party ───────────
        mobile_sql = _mobile_filter(has_crm)
        landline_sql = _landline_filter(has_crm)
        email_sql = _email_filter(has_crm)

        contact_tabs = [
            # (tab_name, party_sql, party_params, color, contact_filter)
            ("Mobile - Republicans",  _REP_SQL, _REP_PARAMS, "8B0000", mobile_sql),
            ("Mobile - Democrats",    _DEM_SQL, _DEM_PARAMS, "1F4E79", mobile_sql),
            ("Mobile - Other",        _OTH_SQL, _OTH_PARAMS, "7B5EA7", mobile_sql),
            ("Landline - Republicans", _REP_SQL, _REP_PARAMS, "8B0000", landline_sql),
            ("Landline - Democrats",   _DEM_SQL, _DEM_PARAMS, "1F4E79", landline_sql),
            ("Landline - Other",       _OTH_SQL, _OTH_PARAMS, "7B5EA7", landline_sql),
            ("Email - Republicans",   _REP_SQL, _REP_PARAMS, "8B0000", email_sql),
            ("Email - Democrats",     _DEM_SQL, _DEM_PARAMS, "1F4E79", email_sql),
            ("Email - Other",         _OTH_SQL, _OTH_PARAMS, "7B5EA7", email_sql),
        ]

        print("\nCreating contact-method tabs...")
        for tab_name, party_sql, party_params, color, contact_filter in contact_tabs:
            print(f"Creating {tab_name} tab...")
            create_full_party_tab(
                wb, conn, district_type, district_number,
                tab_name, party_sql, party_params, color, columns,
                extra_sql=contact_filter,
            )

        # ── Summary tab (row counts — must come after all data tabs) ───────
        print("\nCreating Summary tab...")
        create_contact_summary_tab(wb, conn, district_type, district_number)

        # ── Save ────────────────────────────────────────────────────────────
        output_dir = get_output_dir(district_type, district_number_padded)
        output_file = output_dir / f"{district_type}_{district_number_padded}_Voter_Contact_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print(f"\nSaving workbook...")
        wb.save(output_file)

        print(f"\n{'='*80}")
        print(f"  SUCCESS!")
        print(f"{'='*80}")
        print(f"\nFile saved to:")
        print(f"  {output_file}")

        # Upload to SFTP if configured (Railway / production)
        try:
            sys.path.insert(0, str(Path(__file__).parent.parent.parent))
            from utils_sftp import sftp_upload
            url = sftp_upload(str(output_file), remote_dir="exports")
            if url:
                print(f"\n  Download: {url}")
        except Exception:
            pass

        print(f"\nWorkbook contains:")
        print(f"  - Summary (row counts), BOE Donors, National Donor tabs")
        if has_cfb:
            print(f"  - CFB Donors tab")
        print(f"  - Registered Republicans / Democrats / Other (all columns)")
        print(f"  - Mobile: Republicans / Democrats / Other")
        print(f"  - Landline: Republicans / Democrats / Other")
        print(f"  - Email: Republicans / Democrats / Other")
        print(f"\nTotal tabs: {len(wb.worksheets)}")
        print()

    finally:
        conn.close()


if __name__ == "__main__":
    main()
