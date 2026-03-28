#!/usr/bin/env python3
"""
export/export_statewide.py
==========================
Generate a single statewide summary Excel workbook.

No individual voter rows — only aggregate counts.  Each sheet is a
pivot table: districts (or counties) as rows, audiences as columns.

Sheets produced:
  Summary         — statewide totals per audience + % of all voters
  By LD           — count per audience × Legislative District
  By SD           — count per audience × State Senate District
  By CD           — count per audience × Congressional District
  By County       — count per audience × County

Usage:
    python export/export_statewide.py
    python export/export_statewide.py --out output/Statewide_2026-03.xlsx
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
from utils.db import get_conn

# ── Style constants ───────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
SUB_FILL   = PatternFill("solid", fgColor="2E75B6")
SUB_FONT   = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL   = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FONT = Font(bold=True, size=10)
BODY_FONT  = Font(size=10)
PCT_FMT    = '0.0%'
NUM_FMT    = '#,##0'
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",   vertical="center")


def _hdr(cell, value, fill=HDR_FILL, font=HDR_FONT):
    cell.value     = value
    cell.fill      = fill
    cell.font      = font
    cell.alignment = CENTER


def _autowidth(ws, min_w=8, max_w=40):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value:
                best = min(max(best, len(str(cell.value)) + 2), max_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = best


# ── Data fetchers ─────────────────────────────────────────────────────────────

def fetch_audiences(conn) -> list[str]:
    """Return sorted list of distinct audience names."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT audience
            FROM nys_voter_tagging.voter_audience_bridge
            ORDER BY audience
        """)
        return [r[0] for r in cur.fetchall()]


def fetch_total_voters(conn) -> int:
    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM nys_voter_tagging.voter_file")
        return cur.fetchone()[0]


def fetch_statewide_counts(conn, audiences: list[str]) -> dict[str, int]:
    """Return {audience: statewide_count} for every audience."""
    placeholders = ','.join(['%s'] * len(audiences))
    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT audience, COUNT(*) AS cnt
            FROM nys_voter_tagging.voter_audience_bridge
            WHERE audience IN ({placeholders})
            GROUP BY audience
        """, audiences)
        return {r[0]: r[1] for r in cur.fetchall()}


def fetch_breakdown(conn, col: str) -> dict[tuple, int]:
    """
    Return {(district_value, audience): count} for one district column.
    Uses a single GROUP BY query for all districts at once.
    """
    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT vf.`{col}`, vab.audience, COUNT(*) AS cnt
            FROM nys_voter_tagging.voter_file vf
            INNER JOIN nys_voter_tagging.voter_audience_bridge vab
                ON vf.StateVoterId = vab.StateVoterId
            WHERE vf.`{col}` IS NOT NULL AND vf.`{col}` != ''
            GROUP BY vf.`{col}`, vab.audience
            ORDER BY vf.`{col}`, vab.audience
        """)
        return {(r[0], r[1]): r[2] for r in cur.fetchall()}


def fetch_district_totals(conn, col: str) -> dict[str, int]:
    """Return {district_value: total_voter_count} for one district column."""
    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT `{col}`, COUNT(*) AS cnt
            FROM nys_voter_tagging.voter_file
            WHERE `{col}` IS NOT NULL AND `{col}` != ''
            GROUP BY `{col}`
            ORDER BY `{col}`
        """)
        return {r[0]: r[1] for r in cur.fetchall()}


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_summary_sheet(wb, audiences: list[str], sw_counts: dict, total: int):
    ws = wb.create_sheet("Summary")
    ws.freeze_panes = "A3"

    # Title
    ws["A1"] = f"NYS Voter File — Statewide Audience Summary  ({datetime.now():%B %d, %Y})"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws.merge_cells("A1:D1")

    # Header row
    for col, label in enumerate(["Audience", "Voters", "% of State", "Notes"], 1):
        _hdr(ws.cell(row=2, column=col), label)

    # Data rows
    for i, aud in enumerate(audiences, 3):
        cnt = sw_counts.get(aud, 0)
        pct = cnt / total if total else 0
        ws.cell(row=i, column=1, value=aud).font         = BODY_FONT
        ws.cell(row=i, column=2, value=cnt).number_format = NUM_FMT
        ws.cell(row=i, column=2).font                    = BODY_FONT
        ws.cell(row=i, column=3, value=pct).number_format = PCT_FMT
        ws.cell(row=i, column=3).font                    = BODY_FONT
        if i % 2 == 0:
            for c in range(1, 5):
                ws.cell(row=i, column=c).fill = ALT_FILL

    # Total row
    last = len(audiences) + 3
    ws.cell(row=last, column=1, value="TOTAL REGISTERED VOTERS").font = TOTAL_FONT
    ws.cell(row=last, column=2, value=total).number_format             = NUM_FMT
    ws.cell(row=last, column=2).font                                   = TOTAL_FONT
    ws.cell(row=last, column=3, value=1.0).number_format               = PCT_FMT
    ws.cell(row=last, column=3).font                                   = TOTAL_FONT

    _autowidth(ws)
    ws.column_dimensions["A"].width = 46
    ws.row_dimensions[1].height     = 22


def build_breakdown_sheet(
    wb,
    sheet_name: str,
    col: str,
    audiences: list[str],
    breakdown: dict,
    dist_totals: dict,
):
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "B3"

    n_aud = len(audiences)

    # Row 1: sheet title
    ws.cell(row=1, column=1, value=f"NYS — {sheet_name}  ({datetime.now():%B %d, %Y})")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="1F4E79")
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=n_aud + 2)

    # Row 2: column headers
    _hdr(ws.cell(row=2, column=1), col.replace("Name", ""))
    _hdr(ws.cell(row=2, column=2), "Total Voters")
    for j, aud in enumerate(audiences, 3):
        _hdr(ws.cell(row=2, column=j), aud)

    # Sort districts naturally (numeric where possible)
    def _sort_key(d):
        try:    return (0, int(d))
        except: return (1, d)

    districts = sorted(dist_totals.keys(), key=_sort_key)

    for i, dist in enumerate(districts, 3):
        fill = ALT_FILL if i % 2 == 0 else None
        c = ws.cell(row=i, column=1, value=dist)
        c.font = BODY_FONT
        if fill: c.fill = fill

        tot = dist_totals.get(dist, 0)
        c2 = ws.cell(row=i, column=2, value=tot)
        c2.number_format = NUM_FMT
        c2.font          = BODY_FONT
        if fill: c2.fill = fill

        for j, aud in enumerate(audiences, 3):
            cnt = breakdown.get((dist, aud), 0)
            cell = ws.cell(row=i, column=j, value=cnt)
            cell.number_format = NUM_FMT
            cell.font          = BODY_FONT
            if fill: cell.fill = fill

    # Totals row at the bottom
    last = len(districts) + 3
    ws.cell(row=last, column=1, value="STATEWIDE TOTAL").font = TOTAL_FONT
    grand = sum(dist_totals.values())
    c2 = ws.cell(row=last, column=2, value=grand)
    c2.number_format = NUM_FMT
    c2.font          = TOTAL_FONT
    for j, aud in enumerate(audiences, 3):
        col_total = sum(breakdown.get((d, aud), 0) for d in districts)
        ct = ws.cell(row=last, column=j, value=col_total)
        ct.number_format = NUM_FMT
        ct.font          = TOTAL_FONT

    _autowidth(ws)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.row_dimensions[1].height     = 20
    ws.row_dimensions[2].height     = 40   # wrap audience names


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate statewide audience summary workbook")
    parser.add_argument(
        "--out", metavar="PATH",
        default=None,
        help="Output file path (default: output/Statewide_YYYY-MM-DD.xlsx)",
    )
    args = parser.parse_args()

    out_path = Path(args.out) if args.out else (
        Path(__file__).parent.parent / "output" /
        f"Statewide_{datetime.now():%Y-%m-%d}.xlsx"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*72}")
    print(f"  NYS STATEWIDE AUDIENCE EXPORT")
    print(f"  Output: {out_path}")
    print(f"  {datetime.now():%Y-%m-%d %H:%M:%S}")
    print(f"{'='*72}\n")

    conn = get_conn("nys_voter_tagging", autocommit=True)

    print("  Fetching audience list...")
    audiences = fetch_audiences(conn)
    print(f"  {len(audiences)} audiences found.")

    print("  Counting total voters...")
    total = fetch_total_voters(conn)
    print(f"  {total:,} total voters.")

    print("  Fetching statewide audience counts...")
    sw_counts = fetch_statewide_counts(conn, audiences)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("  Building Summary sheet...")
    build_summary_sheet(wb, audiences, sw_counts, total)

    breakdowns = [
        ("By LD",     "LDName"),
        ("By SD",     "SDName"),
        ("By CD",     "CDName"),
        ("By County", "CountyName"),
    ]

    for sheet_name, col in breakdowns:
        print(f"  Building {sheet_name} sheet...")
        breakdown   = fetch_breakdown(conn, col)
        dist_totals = fetch_district_totals(conn, col)
        build_breakdown_sheet(wb, sheet_name, col, audiences, breakdown, dist_totals)

    conn.close()

    print(f"\n  Saving workbook...")
    wb.save(out_path)
    print(f"\n  Done: {out_path}")
    print(f"  Sheets: Summary + {len(breakdowns)} breakdown tabs")
    print()


if __name__ == "__main__":
    main()
